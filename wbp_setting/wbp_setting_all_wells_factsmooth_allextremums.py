# Скрипт для анализа качества настройки пластового давления по скважинам
# В качестве входных данных - текстовый файл с фактическими давлениями по формату "Скважина	Дата(ДД.ММ.ГГГГ)	Давления(в барах)"

from __future__ import annotations

import os
import warnings
import json
import http.client
from typing import Any, Dict, List, Optional, Tuple, Union
from datetime import datetime

import numpy as np
import pandas as pd
from scipy.signal import savgol_filter, find_peaks
from scipy.interpolate import interp1d
from openpyxl import Workbook


# ============================================================================
# БЛОК, ЗАПОЛНЯЕМЫЙ ПОЛЬЗОВАТЕЛЕМ
WBP_FACT_TXT: str = "scripts_data/Pfkt0_156.inc"  # Относительный путь до файла с фактическими давлениями
MODEL_NAMES: list[str] = [
    "Hist_L0_adapt_GC_AQ_bz_YAMB_SWL_new_swl1.2",
]

# Глобальные настройки сглаживания
SMOOTHING_WINDOW: int = 51  # Размер окна сглаживания (нечетное число)
SMOOTHING_POLYORDER: int = 3  # Порядок полинома для фильтра Савицкого-Голея
SMOOTHING_INTERP_METHOD: str = 'cubic'  # Метод интерполяции для сглаженной кривой
MIN_POINTS_FOR_SMOOTHING: int = 5  # Минимальное количество точек для сглаживания

# Настройки поиска экстремумов для исторических данных
MIN_DISTANCE_DAYS: int = 60  # Минимальное расстояние между экстремумами в днях
PROMINENCE_PERCENT: float = 2.0  # Минимальная значимость экстремума в процентах от среднего значения
MAX_CYCLE_DAYS: int = 400  # Максимальная длина цикла в днях
EDGE_BUFFER_DAYS: int = 30  # Буфер для обработки краев данных в днях (используется для поиска экстремумов в начале и конце)
EXCLUDE_END_DAYS: int = 60  # Количество дней с конца периода, в которых экстремумы будут исключены (для корректного сопоставления факта и моделей)

# Настройки поиска экстремумов для модельных данных (более чувствительные)
MODEL_MIN_DISTANCE_DAYS: int = 45  # Минимальное расстояние между экстремумами в днях для моделей
MODEL_PROMINENCE_PERCENT: float = 1.0  # Минимальная значимость экстремума в процентах для моделей (более чувствительно)
MODEL_MAX_CYCLE_DAYS: int = 400  # Максимальная длина цикла в днях для моделей
MODEL_EDGE_BUFFER_DAYS: int = 30  # Буфер для обработки краев данных в днях для моделей (используется для поиска экстремумов в начале и конце)
MODEL_EXCLUDE_END_DAYS: int = 60  # Количество дней с конца периода, в которых экстремумы будут исключены для моделей

# Настройки сервера для построения графиков
GRAPH_SERVER_HOST: str = "localhost"  # Хост сервера
GRAPH_SERVER_PORT: int = 8000  # Порт сервера
GRAPH_SERVER_ENDPOINT: str = "/api/generate_graphs"  # Эндпойнт для запроса графиков
GRAPH_ARCHIVE_NAME: str = "graphs_archive.zip"  # Имя архива с графиками для сохранения

PROJECT_FOLDER_PATH: str = get_project_folder()  # ВСТРОЕННАЯ В ИНСТРУМЕНТ ФУНКЦИЯ - ВОЗВРАЩАЕТ ПУТЬ К ПРОЕКТУ (пример - I:/L/phg/RedGift_USG)


def parse_fact_well_data(file_path: str) -> pd.DataFrame:
    """
    Парсит файл с данными скважин и возвращает DataFrame.
    
    Parameters
    ----------
    file_path : str
        Путь к файлу с фактическими данными по формату "Скважина Дата(ДД.ММ.ГГГГ) Давления(в барах)"
    
    Returns
    -------
    pd.DataFrame
        DataFrame с колонками: well_fact, date_fact, wpb_bar_fact
    
    Raises
    ------
    ValueError
        Если файл не может быть прочитан ни в одной из поддерживаемых кодировок
    """
    print(f"Загрузка фактических данных из файла: {file_path}")
    
    encodings: List[str] = ['utf-8', 'cp1251', 'cp1252', 'latin1', 'iso-8859-1', 'windows-1251']
    lines: List[str] = []
    
    # Читаем файл
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = f.readlines()
            print(f"  Файл прочитан с кодировкой: {encoding}")
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('Файл не читается в представленных кодировках')
    
    data: List[Dict[str, Any]] = []
    error_count: int = 0
    
    for line in lines:
        # Пропускаем строки с комментариями и служебной информацией
        if line.startswith('--') or line.startswith('PC_') or 'C:\\' in line:
            continue
        
        # Разделяем строку на элементы
        parts: List[str] = line.strip().split()
        
        # Проверяем, что строка содержит 3 значения (well_fact, date_fact, wpb_bar_fact)
        if len(parts) == 3:
            try:
                well_fact: str = str(parts[0])
                date_str: str = parts[1]
                wpb_bar_fact: float = float(parts[2])
                
                # Преобразуем дату в формат datetime
                try:
                    date_fact: datetime = datetime.strptime(date_str, '%d.%m.%Y')
                except ValueError:
                    error_count += 1
                    continue
                
                data.append({
                    'well_fact': well_fact,
                    'date_fact': date_fact,
                    'wpb_bar_fact': wpb_bar_fact
                })
            except (ValueError, IndexError):
                error_count += 1
                continue
    
    # Создаем DataFrame
    df: pd.DataFrame = pd.DataFrame(data)
    
    # Сортируем по well_fact и date_fact для удобства
    if not df.empty:
        df = df.sort_values(['well_fact', 'date_fact']).reset_index(drop=True)
    
    if error_count > 0:
        print(f"  Предупреждение: пропущено {error_count} строк с ошибками")
    
    print(f"  Загружено {len(df)} записей для {df['well_fact'].nunique()} скважин")
    return df


def get_unique_fact_wells(df: pd.DataFrame) -> List[str]:
    """
    Извлекает список уникальных имен скважин из DataFrame.
    
    Parameters
    ----------
    df : pd.DataFrame
        DataFrame с колонкой 'well_fact'
    
    Returns
    -------
    List[str]
        Отсортированный список уникальных имен скважин
    """
    if df is None or df.empty:
        return []
    
    if 'well_fact' not in df.columns:
        return []
    
    # Извлекаем уникальные значения и сортируем их
    unique_wells: List[str] = sorted(df['well_fact'].unique().tolist())
    
    return unique_wells


def get_raw_model_data(
    model_names: List[str], 
    well_names: List[str], 
    parameters: Optional[List[str]] = None
) -> Dict[str, Dict[str, Any]]:
    """
    Получить сырые данные из моделей (без интерполяции).
    
    Parameters
    ----------
    model_names : List[str]
        Список имен моделей для загрузки
    well_names : List[str]
        Список имен скважин для загрузки
    parameters : Optional[List[str]]
        Список параметров для загрузки. По умолчанию: ['wbp', 'wgpr', 'wgir']
    
    Returns
    -------
    Dict[str, Dict[str, Any]]
        Словарь с данными моделей:
        {
            model_name: {
                'dates': List[datetime],
                'well_data': {
                    well_name: {
                        param: List[float]
                    }
                }
            }
        }
    """
    if parameters is None:
        parameters = ['wbp', 'wgpr', 'wgir']  # Давление, добыча газа, закачка газа
    
    models_data: Dict[str, Dict[str, Any]] = {}
    
    for model_name in model_names:
        print(f"Загрузка модели: {model_name}")
        
        try:
            # Получаем объект модели
            model = get_model_by_name(model_name)  # ВСТРОЕННАЯ В ИНСТРУМЕНТ ФУНКЦИЯ - ВОЗВРАЩАЕТ ОБЪЕКТ-МОДЕЛЬ: <class '__main__.tnav.gc.model_class'>
            
            # Получаем все временные шаги
            timesteps = get_all_timesteps()  # ВСТРОЕННАЯ В ИНСТРУМЕНТ ФУНКЦИЯ - ВОЗВРАЩАЕТ список ОБЪЕКТОВ-ВРЕМЕННЫХ ШАГОВ: <class '__main__.tnav.gc.timestep_class'>
            model_dates = [t.to_datetime() for t in timesteps]
            
            # Создаем структуру для данных модели
            model_data: Dict[str, Any] = {
                'dates': model_dates,
                'well_data': {}
            }
            
            # Получаем список всех скважин в модели
            try:
                model_wells = get_all_wells()
                model_well_names: List[str] = [w.name for w in model_wells]
            except Exception:
                model_well_names = []
            
            loaded_wells: int = 0
            skipped_wells: int = 0
            
            # Для каждой запрошенной скважины
            for well_name in well_names:
                # Проверяем, есть ли скважина в модели
                if well_name not in model_well_names:
                    skipped_wells += 1
                    continue
                
                try:
                    well = get_well_by_name(well_name)  # ВСТРОЕННАЯ В ИНСТРУМЕНТ ФУНКЦИЯ - ВОЗВРАЩАЕТ ОБЪЕКТ-СКВАЖИНУ о ее имени: <class '__main__.tnav.gc.well_class'>
                    well_data = {}
                    
                    # Загружаем каждый параметр
                    for param in parameters:
                        try:
                            # Пробуем разные варианты для давления
                            if param == 'wbp':
                                try:
                                    graph_data = wbp[model, well]
                                except Exception:
                                    try:
                                        graph_data = wbhp[model, well]
                                    except Exception:
                                        try:
                                            graph_data = wbhp_h[model, well]
                                        except Exception:
                                            continue
                            elif param == 'wgpr':
                                try:
                                    graph_data = wgpr[model, well]
                                except Exception:
                                    continue
                            elif param == 'wgir':
                                try:
                                    graph_data = wgir[model, well]
                                except Exception:
                                    continue
                            else:
                                continue
                            
                            # Извлекаем значения
                            values: List[float] = []
                            for t in timesteps:
                                try:
                                    value = graph_data[t]
                                    values.append(float(value))
                                except Exception:
                                    values.append(np.nan)
                            
                            well_data[param] = values
                            
                        except Exception:
                            continue
                    
                    if well_data:
                        model_data['well_data'][well_name] = well_data
                        loaded_wells += 1
                    
                except Exception:
                    skipped_wells += 1
                    continue
            
            models_data[model_name] = model_data
            print(f"  Загружено: {loaded_wells} скважин, пропущено: {skipped_wells}")
            
        except Exception as e:
            print(f"  Ошибка при загрузке модели {model_name}: {e}")
            continue
    
    print(f"Итог: загружены данные из {len(models_data)} моделей")
    return models_data


def create_combined_dataframe_per_well_without_interpolation(
    models_raw: Dict[str, Dict[str, Any]],
    historical_df: pd.DataFrame,
    well_column: str = 'well_fact',
    date_column: str = 'date_fact',
    pressure_column: str = 'wpb_bar_fact'
) -> Dict[str, pd.DataFrame]:
    """
    Создать объединенный DataFrame с данными, сгруппированными по скважинам.
    БЕЗ интерполяции - берем фактические даты и модельные даты как есть.
    
    Parameters
    ----------
    models_raw : Dict[str, Dict[str, Any]]
        Сырые данные моделей
    historical_df : pd.DataFrame
        Исторические данные
    well_column : str
        Название колонки со скважинами
    date_column : str
        Название колонки с датами
    pressure_column : str
        Название колонки с давлениями
    
    Returns
    -------
    Dict[str, pd.DataFrame]
        Словарь DataFrame по скважинам
    """
    # Преобразуем исторические даты
    historical_df = historical_df.copy()
    historical_df[date_column] = pd.to_datetime(historical_df[date_column])
    
    well_dataframes: Dict[str, pd.DataFrame] = {}
    
    # Сначала получаем список всех скважин
    all_wells: set = set(historical_df[well_column].unique())
    for model_data in models_raw.values():
        all_wells.update(model_data.get('well_data', {}).keys())
    
    print(f"  Обработка {len(all_wells)} скважин...")
    
    for well in all_wells:
        all_records: List[Dict[str, Any]] = []
        
        # Получаем фактические данные для этой скважины
        well_historical: pd.DataFrame = historical_df[historical_df[well_column] == well].copy()
        
        # Определяем минимальную и максимальную даты для фактических данных
        min_fact_date: Optional[datetime] = None
        max_fact_date: Optional[datetime] = None
        if not well_historical.empty:
            min_fact_date = well_historical[date_column].min()
            max_fact_date = well_historical[date_column].max()
        
        # 1. Добавляем фактические данные для этой скважины (vectorized)
        if not well_historical.empty:
            fact_mask = well_historical[pressure_column].notna()
            fact_filtered = well_historical[fact_mask]
            if not fact_filtered.empty:
                fact_records = [
                    {
                        'date': date,
                        'model': 'HISTORICAL',
                        'parameter': 'pressure',
                        'value': value
                    }
                    for date, value in zip(fact_filtered[date_column], fact_filtered[pressure_column])
                ]
                all_records.extend(fact_records)
        
        # 2. Добавляем модельные данные для этой скважины (без интерполяции)
        for model_name, model_info in models_raw.items():
            if well in model_info.get('well_data', {}):
                model_dates: List[datetime] = model_info['dates']
                well_data: Dict[str, List[float]] = model_info['well_data'][well]
                
                # Фильтруем модельные даты по диапазону фактических дат
                if min_fact_date and max_fact_date:
                    date_series: pd.Series = pd.Series(model_dates)
                    date_mask: pd.Series = (date_series >= min_fact_date) & (date_series <= max_fact_date)
                    filtered_indices: np.ndarray = np.where(date_mask)[0]
                else:
                    filtered_indices = np.arange(len(model_dates))
                
                if len(filtered_indices) == 0:
                    continue
                
                # Для каждого параметра
                for param, all_values in well_data.items():
                    if len(all_values) != len(model_dates):
                        continue
                    
                    # Сопоставляем имена параметров
                    param_display: str = {
                        'wbp': 'pressure',
                        'wgpr': 'gas_rate',
                        'wgir': 'gas_injection'
                    }.get(param, param)
                    
                    # Добавляем записи для отфильтрованных дат (vectorized)
                    for idx in filtered_indices:
                        date = model_dates[idx]
                        value = all_values[idx]
                        
                        # Пропускаем NaN значения
                        if pd.isna(value):
                            continue
                            
                        all_records.append({
                            'date': date,
                            'model': model_name,
                            'parameter': param_display,
                            'value': value
                        })
        
        # Создаем DataFrame для скважины
        if all_records:
            df_well: pd.DataFrame = pd.DataFrame(all_records)
            df_well = df_well.sort_values(['date', 'parameter', 'model']).reset_index(drop=True)
            well_dataframes[well] = df_well
    
    print(f"  Создано DataFrame для {len(well_dataframes)} скважин")
    return well_dataframes


def get_unified_data_per_well_without_interpolation(
    model_names: List[str],
    historical_df: pd.DataFrame,
    well_column: str = 'well_fact',
    date_column: str = 'date_fact',
    pressure_column: str = 'wpb_bar_fact'
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, Any]]]:
    """
    Основная функция: получить унифицированные данные, сгруппированные по скважинам.
    БЕЗ интерполяции - берем фактические даты и модельные даты как есть.
    
    Parameters
    ----------
    model_names : List[str]
        Список имен моделей
    historical_df : pd.DataFrame
        Исторические данные
    well_column : str
        Название колонки со скважинами
    date_column : str
        Название колонки с датами
    pressure_column : str
        Название колонки с давлениями
    
    Returns
    -------
    Tuple[Dict[str, pd.DataFrame], Dict[str, Dict[str, Any]]]
        (well_dataframes, models_raw)
    """
    print("Получение унифицированных данных по скважинам...")
    
    # 1. Получаем список скважин из фактических данных
    well_names: List[str] = get_unique_fact_wells(historical_df)
    if not well_names:
        print("Ошибка: не удалось получить список скважин из фактических данных")
        return {}, {}
    
    print(f"  Скважин для анализа: {len(well_names)}")
    
    # 2. Загружаем сырые данные из моделей
    models_raw: Dict[str, Dict[str, Any]] = get_raw_model_data(
        model_names, well_names, ['wbp', 'wgpr', 'wgir']
    )
    
    if not models_raw:
        print("Ошибка: не удалось загрузить данные моделей")
        return {}, {}
    
    # 3. Создаем объединенные DataFrame по скважинам (без интерполяции)
    well_dataframes: Dict[str, pd.DataFrame] = create_combined_dataframe_per_well_without_interpolation(
        models_raw, historical_df,
        well_column=well_column,
        date_column=date_column,
        pressure_column=pressure_column
    )
    
    # 4. Анализ загруженных параметров
    param_stats: Dict[str, set] = {}
    for well, df_well in well_dataframes.items():
        params = df_well['parameter'].unique()
        for param in params:
            if param not in param_stats:
                param_stats[param] = set()
            param_stats[param].add(well)
    
    print(f"  Параметры: {', '.join(f'{p}({len(ws)} скв.)' for p, ws in param_stats.items())}")
    
    return well_dataframes, models_raw


def smooth_pressure_timeseries(df: pd.DataFrame) -> pd.DataFrame:
    """
    Создает сглаженную кривую для временного ряда давления.
    Сглаживание применяется только к существующим данным, но значения
    сглаженной кривой вычисляются для всех дат.
    
    Parameters
    ----------
    df : pd.DataFrame
        Датафрейм с колонками 'date' (строки в формате ДД.ММ.ГГГГ) и 'pressure_fact'
    
    Returns
    -------
    pd.DataFrame
        Исходный датафрейм с добавленной колонкой 'pressure_smoothed'
    """
    # Создаем копию датафрейма
    df_result = df.copy()
    
    # Преобразуем даты в datetime
    df_result['date_dt'] = pd.to_datetime(df_result['date'], format='%d.%m.%Y')
    
    # Сортируем по дате
    df_result = df_result.sort_values('date_dt').reset_index(drop=True)
    
    # Шаг 1: Получаем только фактические данные (без NaN)
    fact_mask = df_result['pressure_fact'].notna()
    fact_dates = df_result.loc[fact_mask, 'date_dt']
    fact_pressures = df_result.loc[fact_mask, 'pressure_fact'].values
    
    # Проверяем, достаточно ли точек для сглаживания
    if len(fact_pressures) < MIN_POINTS_FOR_SMOOTHING:
        warnings.warn(f"Недостаточно точек для сглаживания. Доступно: {len(fact_pressures)}, требуется: {MIN_POINTS_FOR_SMOOTHING}")
        # Если точек недостаточно, используем простую линейную интерполяцию
        df_result['pressure_smoothed'] = np.nan
        # Интерполируем только между фактическими точками
        if len(fact_pressures) >= 2:
            # Создаем временную шкалу в днях от первой даты
            all_dates_numeric = (df_result['date_dt'] - df_result['date_dt'].min()).dt.days.values
            fact_dates_numeric = (fact_dates - df_result['date_dt'].min()).dt.days.values
            
            # Линейная интерполяция
            interp_func = interp1d(fact_dates_numeric, fact_pressures, 
                                  kind='linear', bounds_error=False, 
                                  fill_value='extrapolate')
            
            # Получаем значения для всех дат
            smoothed_all = interp_func(all_dates_numeric)
            df_result['pressure_smoothed'] = smoothed_all
        
        df_result.drop(columns=['date_dt'], inplace=True)
        return df_result
    
    # Шаг 2: Создаем числовую временную шкалу для фактических точек
    # Используем дни от первой фактической даты
    first_fact_date = fact_dates.iloc[0]
    fact_dates_numeric = (fact_dates - first_fact_date).dt.days.values
    
    # Шаг 3: Создаем равномерную временную шкалу для сглаживания
    # (фильтр Савицкого-Голея требует равномерной сетки)
    min_date_num = fact_dates_numeric.min()
    max_date_num = fact_dates_numeric.max()
    
    # Создаем равномерную сетку с шагом 1 день
    uniform_dates_num = np.arange(min_date_num, max_date_num + 1)
    
    # Интерполируем фактические данные на равномерную сетку
    # (только для целей сглаживания)
    interp_func_uniform = interp1d(fact_dates_numeric, fact_pressures, 
                                  kind=SMOOTHING_INTERP_METHOD, 
                                  bounds_error=False, 
                                  fill_value='extrapolate')
    
    pressures_uniform = interp_func_uniform(uniform_dates_num)
    
    # Шаг 4: Применяем фильтр Савицкого-Голея к равномерной сетке
    # Настраиваем размер окна
    window_length = min(SMOOTHING_WINDOW, len(pressures_uniform))
    if window_length % 2 == 0:
        window_length -= 1  # Делаем нечетным
        if window_length < 3:
            window_length = 3
    
    # Убедимся, что порядок полинома меньше размера окна
    polyorder = min(SMOOTHING_POLYORDER, window_length - 1)
    
    try:
        smoothed_uniform = savgol_filter(pressures_uniform, 
                                        window_length=window_length,
                                        polyorder=polyorder,
                                        mode='interp')
    except Exception as e:
        warnings.warn(f"Ошибка при сглаживании: {str(e)}. Использую интерполированные значения.")
        smoothed_uniform = pressures_uniform
    
    # Шаг 5: Создаем интерполяционную функцию для сглаженной кривой
    # (снова используем фактическую временную шкалу)
    # Создаем функцию, которая интерполирует сглаженные значения
    # на основе равномерной сетки
    interp_smoothed = interp1d(uniform_dates_num, smoothed_uniform,
                              kind=SMOOTHING_INTERP_METHOD,
                              bounds_error=False,
                              fill_value='extrapolate')
    
    # Шаг 6: Вычисляем сглаженные значения для всех дат (включая пропуски)
    all_dates_numeric = (df_result['date_dt'] - first_fact_date).dt.days.values
    smoothed_all = interp_smoothed(all_dates_numeric)
    
    # Ограничиваем значения разумными пределами
    min_pressure = max(0, np.nanmin(fact_pressures) * 0.5)
    max_pressure = np.nanmax(fact_pressures) * 1.5
    smoothed_all = np.clip(smoothed_all, min_pressure, max_pressure)
    
    df_result['pressure_smoothed'] = smoothed_all
    
    # Шаг 7: Убедимся, что сглаженная кривая не имеет NaN
    # (интерполяция с fill_value='extrapolate' должна это обеспечить)
    if df_result['pressure_smoothed'].isna().any():
        # Заполняем оставшиеся NaN линейной интерполяцией
        df_result['pressure_smoothed'] = df_result['pressure_smoothed'].interpolate(
            method='linear', limit_direction='both'
        )
    
    # Удаляем вспомогательные колонки
    df_result.drop(columns=['date_dt'], inplace=True)
    
    return df_result


def find_extremes_improved_v2(
    df: pd.DataFrame,
    min_distance_days: int = MIN_DISTANCE_DAYS,
    prominence_percent: float = PROMINENCE_PERCENT,
    max_cycle_days: int = MAX_CYCLE_DAYS,
    edge_buffer_days: int = EDGE_BUFFER_DAYS,
    exclude_end_days: int = 0
) -> pd.DataFrame:
    """
    Улучшенный алгоритм поиска экстремумов для циклических данных с годовыми циклами.
    
    Parameters
    ----------
    df : pd.DataFrame
        Датафрейм с колонками ['date', 'pressure_smoothed']
    min_distance_days : int
        Минимальное расстояние между экстремумами в днях (по умолчанию 60)
    prominence_percent : float
        Минимальная значимость экстремума в процентах от среднего значения (по умолчанию 2%)
    max_cycle_days : int
        Максимальная длина цикла в днях (по умолчанию 400)
    edge_buffer_days : int
        Буфер для обработки краев данных в днях (по умолчанию 30)
    exclude_end_days : int
        Количество дней с конца периода, в которых экстремумы будут исключены (по умолчанию 0)
        Используется для корректного сопоставления факта и моделей, когда в конце периода
        экстремумы могут быть определены только в одной из серий
    
    Returns
    -------
    pd.DataFrame
        Датафрейм с добавленными колонками 'maxima' и 'minima'
    """
    # Копируем датафрейм
    result_df: pd.DataFrame = df.copy()
    
    # Преобразуем дату
    if not pd.api.types.is_datetime64_any_dtype(result_df['date']):
        try:
            result_df['date'] = pd.to_datetime(result_df['date'], format='%d.%m.%Y')
        except Exception:
            result_df['date'] = pd.to_datetime(result_df['date'])
    
    # Сортируем по дате
    result_df = result_df.sort_values('date').reset_index(drop=True)
    
    # Инициализируем колонки
    result_df['maxima'] = np.nan
    result_df['minima'] = np.nan
    
    # Получаем значения
    pressures: np.ndarray = result_df['pressure_smoothed'].values
    dates: np.ndarray = result_df['date'].values
    n: int = len(pressures)
    
    if n < 10:
        return result_df
    
    # Вычисляем среднее значение для определения значимости
    avg_pressure: float = np.nanmean(pressures)
    min_prominence: float = avg_pressure * (prominence_percent / 100)
    
    # Конвертируем расстояние в днях в индексы
    if n > 1:
        avg_days_between_points: float = (dates[-1] - dates[0]).astype('timedelta64[D]').astype(int) / (n - 1)
        min_distance_points: int = max(5, int(min_distance_days / avg_days_between_points))
    else:
        min_distance_points = 30
    
    # Инициализируем списки для экстремумов
    all_maxima_indices: List[int] = []
    all_minima_indices: List[int] = []
    
    # 1. Поиск с помощью scipy.signal.find_peaks (основной метод)
    try:
        # Находим максимумы
        max_peaks, _ = find_peaks(
            pressures,
            distance=min_distance_points,
            prominence=min_prominence,
            width=min_distance_points//3,
            rel_height=0.5
        )
        
        # Находим минимумы (инвертируем сигнал)
        min_peaks, _ = find_peaks(
            -pressures,
            distance=min_distance_points,
            prominence=min_prominence,
            width=min_distance_points//3,
            rel_height=0.5
        )
        
        all_maxima_indices = list(max_peaks)
        all_minima_indices = list(min_peaks)
        
    except Exception:
        all_maxima_indices = []
        all_minima_indices = []
    
    # 2. Дополнительный поиск по годовым циклам
    years: np.ndarray = result_df['date'].dt.year.unique()
    yearly_extremes: List[Tuple[int, str, float]] = []
    
    for year in years:
        year_mask: pd.Series = result_df['date'].dt.year == year
        year_indices: np.ndarray = np.where(year_mask)[0]
        
        if len(year_indices) > 30:  # Минимум 30 точек в году
            year_pressures: np.ndarray = pressures[year_indices]
            
            # Находим максимум года
            year_max_idx_local: int = np.argmax(year_pressures)
            year_max_idx: int = year_indices[year_max_idx_local]
            year_max_val: float = year_pressures[year_max_idx_local]
            
            # Находим минимум года
            year_min_idx_local: int = np.argmin(year_pressures)
            year_min_idx: int = year_indices[year_min_idx_local]
            year_min_val: float = year_pressures[year_min_idx_local]
            
            # Проверяем, что это действительно экстремум в окрестности
            window_size: int = min(50, len(year_indices)//3)
            
            # Для максимума
            if year_max_idx_local >= window_size and year_max_idx_local < len(year_indices) - window_size:
                window: np.ndarray = year_pressures[year_max_idx_local-window_size:year_max_idx_local+window_size+1]
                if year_max_val == np.max(window):
                    yearly_extremes.append((year_max_idx, 'max', year_max_val))
            
            # Для минимума
            if year_min_idx_local >= window_size and year_min_idx_local < len(year_indices) - window_size:
                window = year_pressures[year_min_idx_local-window_size:year_min_idx_local+window_size+1]
                if year_min_val == np.min(window):
                    yearly_extremes.append((year_min_idx, 'min', year_min_val))
    
    # 2.5. Дополнительный поиск по полугодовым циклам (каждые 6 месяцев)
    # Это помогает находить экстремумы, которые происходят дважды в год
    semi_annual_extremes: List[Tuple[int, str, float]] = []
    
    # Разбиваем данные на полугодовые периоды
    if n > 60:  # Минимум 60 точек для полугодового анализа
        # Создаем периоды по 6 месяцев
        date_series: pd.Series = pd.Series(result_df['date'])
        date_series = pd.to_datetime(date_series)
        
        # Группируем по полугодиям (январь-июнь и июль-декабрь)
        for year in years:
            # Первое полугодие (январь-июнь)
            h1_mask: pd.Series = (result_df['date'].dt.year == year) & (result_df['date'].dt.month <= 6)
            h1_indices: np.ndarray = np.where(h1_mask)[0]
            
            if len(h1_indices) > 15:  # Минимум 15 точек в полугодии
                h1_pressures: np.ndarray = pressures[h1_indices]
                
                # Максимум первого полугодия
                h1_max_idx_local: int = np.argmax(h1_pressures)
                h1_max_idx: int = h1_indices[h1_max_idx_local]
                h1_max_val: float = h1_pressures[h1_max_idx_local]
                
                # Минимум первого полугодия
                h1_min_idx_local: int = np.argmin(h1_pressures)
                h1_min_idx: int = h1_indices[h1_min_idx_local]
                h1_min_val: float = h1_pressures[h1_min_idx_local]
                
                # Проверяем значимость
                window_size_h1: int = min(30, len(h1_indices)//3)
                if h1_max_idx_local >= window_size_h1 and h1_max_idx_local < len(h1_indices) - window_size_h1:
                    window_h1_max: np.ndarray = h1_pressures[h1_max_idx_local-window_size_h1:h1_max_idx_local+window_size_h1+1]
                    if h1_max_val == np.max(window_h1_max) and h1_max_val - h1_min_val > min_prominence * 0.5:
                        semi_annual_extremes.append((h1_max_idx, 'max', h1_max_val))
                
                if h1_min_idx_local >= window_size_h1 and h1_min_idx_local < len(h1_indices) - window_size_h1:
                    window_h1_min: np.ndarray = h1_pressures[h1_min_idx_local-window_size_h1:h1_min_idx_local+window_size_h1+1]
                    if h1_min_val == np.min(window_h1_min) and h1_max_val - h1_min_val > min_prominence * 0.5:
                        semi_annual_extremes.append((h1_min_idx, 'min', h1_min_val))
            
            # Второе полугодие (июль-декабрь)
            h2_mask: pd.Series = (result_df['date'].dt.year == year) & (result_df['date'].dt.month > 6)
            h2_indices: np.ndarray = np.where(h2_mask)[0]
            
            if len(h2_indices) > 15:  # Минимум 15 точек в полугодии
                h2_pressures: np.ndarray = pressures[h2_indices]
                
                # Максимум второго полугодия
                h2_max_idx_local: int = np.argmax(h2_pressures)
                h2_max_idx: int = h2_indices[h2_max_idx_local]
                h2_max_val: float = h2_pressures[h2_max_idx_local]
                
                # Минимум второго полугодия
                h2_min_idx_local: int = np.argmin(h2_pressures)
                h2_min_idx: int = h2_indices[h2_min_idx_local]
                h2_min_val: float = h2_pressures[h2_min_idx_local]
                
                # Проверяем значимость
                window_size_h2: int = min(30, len(h2_indices)//3)
                if h2_max_idx_local >= window_size_h2 and h2_max_idx_local < len(h2_indices) - window_size_h2:
                    window_h2_max: np.ndarray = h2_pressures[h2_max_idx_local-window_size_h2:h2_max_idx_local+window_size_h2+1]
                    if h2_max_val == np.max(window_h2_max) and h2_max_val - h2_min_val > min_prominence * 0.5:
                        semi_annual_extremes.append((h2_max_idx, 'max', h2_max_val))
                
                if h2_min_idx_local >= window_size_h2 and h2_min_idx_local < len(h2_indices) - window_size_h2:
                    window_h2_min: np.ndarray = h2_pressures[h2_min_idx_local-window_size_h2:h2_min_idx_local+window_size_h2+1]
                    if h2_min_val == np.min(window_h2_min) and h2_max_val - h2_min_val > min_prominence * 0.5:
                        semi_annual_extremes.append((h2_min_idx, 'min', h2_min_val))
    
    # 3. Проверка краевых точек
    edge_extremes = []
    
    # Проверяем первые edge_buffer_days дней
    edge_points = int(edge_buffer_days / avg_days_between_points) if avg_days_between_points > 0 else 30
    edge_points = min(edge_points, n//4)
    
    if edge_points > 5:
        # Проверяем начало данных
        start_window = pressures[:edge_points*2]
        if len(start_window) > 0:
            start_max_idx = np.argmax(start_window)
            start_max_val = start_window[start_max_idx]
            start_min_idx = np.argmin(start_window)
            start_min_val = start_window[start_min_idx]
            
            # Проверяем значимость
            if start_max_idx > 0 and start_max_idx < len(start_window)-1:
                if start_max_val - start_min_val > min_prominence:
                    edge_extremes.append((start_max_idx, 'max', start_max_val))
                    edge_extremes.append((start_min_idx, 'min', start_min_val))
        
        # Проверяем конец данных
        end_window = pressures[-edge_points*2:]
        if len(end_window) > 0:
            end_max_idx = n - len(end_window) + np.argmax(end_window)
            end_max_val = pressures[end_max_idx]
            end_min_idx = n - len(end_window) + np.argmin(end_window)
            end_min_val = pressures[end_min_idx]
            
            if end_max_idx > n - len(end_window) and end_max_idx < n-1:
                if end_max_val - end_min_val > min_prominence:
                    edge_extremes.append((end_max_idx, 'max', end_max_val))
                    edge_extremes.append((end_min_idx, 'min', end_min_val))
    
    # 4. Объединяем все найденные экстремумы
    all_extrema_dict: Dict[str, List[Tuple[int, float]]] = {'max': [], 'min': []}
    
    # Добавляем экстремумы из find_peaks
    for idx in all_maxima_indices:
        all_extrema_dict['max'].append((idx, pressures[idx]))
    for idx in all_minima_indices:
        all_extrema_dict['min'].append((idx, pressures[idx]))
    
    # Добавляем годовые экстремумы
    for idx, typ, val in yearly_extremes:
        all_extrema_dict[typ].append((idx, val))
    
    # Добавляем полугодовые экстремумы
    for idx, typ, val in semi_annual_extremes:
        all_extrema_dict[typ].append((idx, val))
    
    # Добавляем краевые экстремумы
    for idx, typ, val in edge_extremes:
        all_extrema_dict[typ].append((idx, val))
    
    # Удаляем дубликаты и сортируем
    for typ in ['max', 'min']:
        if all_extrema_dict[typ]:
            # Удаляем дубликаты по индексу
            unique_dict: Dict[int, float] = {}
            for idx, val in all_extrema_dict[typ]:
                if idx not in unique_dict:
                    unique_dict[idx] = val
                elif typ == 'max' and val > unique_dict[idx]:
                    unique_dict[idx] = val
                elif typ == 'min' and val < unique_dict[idx]:
                    unique_dict[idx] = val
            
            # Сортируем по индексу
            all_extrema_dict[typ] = sorted([(idx, val) for idx, val in unique_dict.items()])
    
    # 5. Фильтрация и чередование экстремумов
    filtered_maxima = []
    filtered_minima = []
    
    # Объединяем все экстремумы в один отсортированный список
    combined_extrema = []
    for idx, val in all_extrema_dict['max']:
        combined_extrema.append((idx, 'max', val))
    for idx, val in all_extrema_dict['min']:
        combined_extrema.append((idx, 'min', val))
    
    combined_extrema.sort(key=lambda x: x[0])
    
    # Алгоритм чередования с допущениями
    i = 0
    last_type = None
    last_idx = -min_distance_points * 2
    
    while i < len(combined_extrema):
        idx, typ, val = combined_extrema[i]
        
        # Проверяем расстояние до предыдущего экстремума
        if idx - last_idx < min_distance_points:
            # Если слишком близко, выбираем более значимый
            if last_type == 'max' and typ == 'max':
                # Два максимума рядом - выбираем больший
                if val > pressures[last_idx]:
                    # Удаляем предыдущий, добавляем текущий
                    if last_idx in filtered_maxima:
                        filtered_maxima.remove(last_idx)
                    filtered_maxima.append(idx)
                    last_idx = idx
                # Иначе пропускаем текущий
            elif last_type == 'min' and typ == 'min':
                # Два минимума рядом - выбираем меньший
                if val < pressures[last_idx]:
                    if last_idx in filtered_minima:
                        filtered_minima.remove(last_idx)
                    filtered_minima.append(idx)
                    last_idx = idx
            i += 1
            continue
        
        # Проверяем чередование
        if last_type is None or typ != last_type:
            # Если это первый экстремум или типы чередуются
            if typ == 'max':
                filtered_maxima.append(idx)
            else:
                filtered_minima.append(idx)
            
            last_type = typ
            last_idx = idx
            i += 1
        else:
            # Если типы не чередуются, проверяем следующий экстремум
            # Ищем ближайший экстремум другого типа в пределах max_cycle_days
            found_alternate = False
            max_search = min(i + 20, len(combined_extrema))
            
            for j in range(i + 1, max_search):
                idx2, typ2, val2 = combined_extrema[j]
                
                # Проверяем расстояние в днях
                days_diff = (dates[idx2] - dates[idx]).astype('timedelta64[D]').astype(int)
                
                if typ2 != typ and 30 < days_diff < max_cycle_days:
                    # Нашли чередующийся экстремум
                    if typ2 == 'max':
                        filtered_maxima.append(idx2)
                    else:
                        filtered_minima.append(idx2)
                    
                    last_type = typ2
                    last_idx = idx2
                    i = j + 1
                    found_alternate = True
                    break
            
            if not found_alternate:
                # Если не нашли чередующийся, пропускаем текущий
                i += 1
    
    # 6. Дополнительная проверка пропущенных экстремумов
    # Ищем крупные пропуски между экстремумами
    all_filtered = sorted([(idx, 'max', pressures[idx]) for idx in filtered_maxima] + 
                         [(idx, 'min', pressures[idx]) for idx in filtered_minima])
    
    for k in range(len(all_filtered) - 1):
        idx1, typ1, val1 = all_filtered[k]
        idx2, typ2, val2 = all_filtered[k + 1]
        
        # Вычисляем расстояние в днях
        days_diff = (dates[idx2] - dates[idx1]).astype('timedelta64[D]').astype(int)
        
        # Если большой пропуск (> 250 дней), ищем экстремум в середине
        if days_diff > 250 and typ1 != typ2:
            mid_idx = (idx1 + idx2) // 2
            search_start = max(0, mid_idx - min_distance_points)
            search_end = min(n, mid_idx + min_distance_points)
            
            if search_end - search_start > 10:
                if typ1 == 'max':
                    # Между максимумом и минимумом должен быть минимум
                    search_min_idx = search_start + np.argmin(pressures[search_start:search_end])
                    search_min_val = pressures[search_min_idx]
                    
                    # Проверяем значимость
                    if val1 - search_min_val > min_prominence and val2 - search_min_val > min_prominence:
                        if search_min_idx not in filtered_minima:
                            filtered_minima.append(search_min_idx)
                else:
                    # Между минимумом и максимумом должен быть максимум
                    search_max_idx = search_start + np.argmax(pressures[search_start:search_end])
                    search_max_val = pressures[search_max_idx]
                    
                    # Проверяем значимость
                    if search_max_val - val1 > min_prominence and search_max_val - val2 > min_prominence:
                        if search_max_idx not in filtered_maxima:
                            filtered_maxima.append(search_max_idx)
    
    # 7. Сортировка и удаление дубликатов
    filtered_maxima: List[int] = sorted(list(set(filtered_maxima)))
    filtered_minima: List[int] = sorted(list(set(filtered_minima)))
    
    # 7.5. Исключаем экстремумы в конце периода (если указано)
    if exclude_end_days > 0 and n > 1:
        # Вычисляем дату, до которой нужно исключить экстремумы
        exclude_end_date = dates[-1] - pd.Timedelta(days=exclude_end_days)
        
        # Находим индекс последней даты, которая меньше exclude_end_date
        # Используем бинарный поиск для эффективности
        exclude_end_idx = np.searchsorted(dates, exclude_end_date, side='right')
        
        # Фильтруем экстремумы, которые находятся после exclude_end_idx
        filtered_maxima = [idx for idx in filtered_maxima if idx < exclude_end_idx]
        filtered_minima = [idx for idx in filtered_minima if idx < exclude_end_idx]
    
    # 8. Заполняем датафрейм
    for idx in filtered_maxima:
        if 0 <= idx < n:
            result_df.loc[idx, 'maxima'] = pressures[idx]
    
    for idx in filtered_minima:
        if 0 <= idx < n:
            result_df.loc[idx, 'minima'] = pressures[idx]
    
    return result_df


def compute_smoothed_pressures_and_extremes(
    well_dataframes: Dict[str, pd.DataFrame],
    historical_df: pd.DataFrame
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, Dict[str, float]]]]:
    """
    Вычисляет сглаженные значения давления и экстремумы для каждой скважины.
    
    Parameters
    ----------
    well_dataframes : Dict[str, pd.DataFrame]
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные с фактическими давлениями
        
    Returns
    -------
    Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, Dict[str, float]]]]
        (smoothed_pressures, extremes_data)
        smoothed_pressures: словарь с ключами - имена скважин, значения - словари с датами и сглаженными значениями
        extremes_data: словарь с ключами - имена скважин, значения - словари с датами и экстремумами
    """
    smoothed_pressures: Dict[str, Dict[str, float]] = {}
    extremes_data: Dict[str, Dict[str, Dict[str, float]]] = {}
    
    print(f"  Вычисление сглаженных давлений и экстремумов для {len(well_dataframes)} скважин...")
    
    # Предварительно группируем исторические данные по скважинам для быстрого доступа
    historical_by_well: Dict[str, pd.DataFrame] = {}
    for well_name in historical_df['well_fact'].unique():
        well_data = historical_df[historical_df['well_fact'] == well_name].copy()
        well_data['date_fact_dt'] = pd.to_datetime(well_data['date_fact'])
        # Создаем индекс для быстрого поиска по дате
        historical_by_well[well_name] = well_data.set_index('date_fact_dt')
    
    for well_name, df_well in well_dataframes.items():
        # Получаем предварительно сгруппированные исторические данные
        historical_data_well = historical_by_well.get(well_name)
        
        if historical_data_well is None or historical_data_well.empty:
            smoothed_pressures[well_name] = {}
            extremes_data[well_name] = {'maxima': {}, 'minima': {}}
            continue
        
        # Получаем все уникальные даты для этой скважины из well_dataframes
        all_dates_for_well: np.ndarray = df_well['date'].unique()
        
        if len(all_dates_for_well) == 0:
            smoothed_pressures[well_name] = {}
            extremes_data[well_name] = {'maxima': {}, 'minima': {}}
            continue
        
        # Создаем DataFrame с ВСЕМИ датами (не только историческими)
        # в формате, ожидаемом функцией smooth_pressure_timeseries
        pressure_data: List[Dict[str, Any]] = []
        
        for date in all_dates_for_well:
            # Приводим все даты к единому формату datetime
            if isinstance(date, (pd.Timestamp, datetime)):
                date_dt: Union[pd.Timestamp, datetime] = date
            elif isinstance(date, np.datetime64):
                date_dt = pd.Timestamp(date)
            else:
                try:
                    date_dt = pd.to_datetime(date)
                except Exception:
                    continue
            
            # Ищем историческое давление для этой даты (быстрый доступ через индекс)
            hist_pressure: Optional[float] = None
            
            # Пробуем точное совпадение через индекс
            if date_dt in historical_data_well.index:
                hist_pressure = historical_data_well.loc[date_dt, 'wpb_bar_fact']
            else:
                # Ищем ближайшую дату (в пределах 1 дня) - оптимизированная версия
                # Используем searchsorted для бинарного поиска (быстрее чем полный перебор)
                date_series = historical_data_well.index
                if len(date_series) > 0:
                    # Бинарный поиск ближайшей даты
                    pos = date_series.searchsorted(date_dt)
                    candidates = []
                    
                    # Проверяем позицию и соседние
                    if pos < len(date_series):
                        candidates.append((abs((date_series[pos] - date_dt).days), pos))
                    if pos > 0:
                        candidates.append((abs((date_series[pos-1] - date_dt).days), pos-1))
                    
                    if candidates:
                        min_diff_days, min_idx = min(candidates, key=lambda x: x[0])
                        if min_diff_days <= 1:
                            hist_pressure = historical_data_well.iloc[min_idx]['wpb_bar_fact']
            
            # Сохраняем дату в строковом формате для функции smooth_pressure_timeseries
            pressure_data.append({
                'date': date_dt.strftime('%d.%m.%Y'),
                'pressure_fact': hist_pressure
            })
        
        if not pressure_data:
            smoothed_pressures[well_name] = {}
            extremes_data[well_name] = {'maxima': {}, 'minima': {}}
            continue
        
        # Создаем DataFrame для сглаживания
        pressure_df: pd.DataFrame = pd.DataFrame(pressure_data)
        
        # Применяем сглаживание
        try:
            smoothed_df: pd.DataFrame = smooth_pressure_timeseries(pressure_df)
            
            # Поиск экстремумов в сглаженных данных
            extremes_df: pd.DataFrame = find_extremes_improved_v2(
                smoothed_df,
                exclude_end_days=EXCLUDE_END_DAYS
            )
            
            # Создаем словарь для быстрого доступа к сглаженным значениям по дате (vectorized)
            def normalize_date_key(date_val):
                """Нормализует дату в формат YYYY-MM-DD"""
                try:
                    if isinstance(date_val, str):
                        try:
                            date_dt = datetime.strptime(date_val, '%d.%m.%Y')
                        except Exception:
                            date_dt = pd.to_datetime(date_val)
                    elif isinstance(date_val, (pd.Timestamp, datetime)):
                        date_dt = date_val
                    elif isinstance(date_val, np.datetime64):
                        date_dt = pd.Timestamp(date_val)
                    else:
                        date_dt = pd.to_datetime(date_val)
                    return date_dt.strftime('%Y-%m-%d')
                except Exception:
                    return str(date_val)
            
            # Преобразуем даты в нормализованный формат (vectorized)
            extremes_df['date_key'] = extremes_df['date'].apply(normalize_date_key)
            
            # Создаем словари (vectorized)
            smoothed_dict = dict(zip(
                extremes_df['date_key'],
                extremes_df['pressure_smoothed']
            ))
            
            # Фильтруем и создаем словари экстремумов (vectorized)
            maxima_mask = extremes_df['maxima'].notna()
            minima_mask = extremes_df['minima'].notna()
            
            maxima_dict = dict(zip(
                extremes_df.loc[maxima_mask, 'date_key'],
                extremes_df.loc[maxima_mask, 'maxima']
            )) if maxima_mask.any() else {}
            
            minima_dict = dict(zip(
                extremes_df.loc[minima_mask, 'date_key'],
                extremes_df.loc[minima_mask, 'minima']
            )) if minima_mask.any() else {}
            
            smoothed_pressures[well_name] = smoothed_dict
            extremes_data[well_name] = {
                'maxima': maxima_dict,
                'minima': minima_dict
            }
            
        except Exception as e:
            print(f"  Ошибка при обработке скважины {well_name}: {str(e)}")
            smoothed_pressures[well_name] = {}
            extremes_data[well_name] = {'maxima': {}, 'minima': {}}
    
    print(f"  Обработано {len([w for w in smoothed_pressures if smoothed_pressures[w]])} скважин")
    return smoothed_pressures, extremes_data


def compute_model_extremes(
    well_dataframes: Dict[str, pd.DataFrame]
) -> Dict[str, Dict[str, Dict[str, Dict[str, float]]]]:
    """
    Вычисляет экстремумы для модельных давлений каждой скважины.
    
    Parameters
    ----------
    well_dataframes : Dict[str, pd.DataFrame]
        DataFrame по скважинам
        
    Returns
    -------
    Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Словарь с экстремумами модельных давлений:
        model_extremes[well_name][model_name] = {'maxima': {date: value}, 'minima': {date: value}}
    """
    print("  Вычисление экстремумов для модельных давлений...")
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]] = {}
    
    for well_name, df_well in well_dataframes.items():
        # Получаем список моделей для этой скважины
        models_for_well: np.ndarray = df_well['model'].unique()
        models_for_well = [m for m in models_for_well if m != 'HISTORICAL']
        
        if not models_for_well:
            model_extremes[well_name] = {}
            continue
        
        model_extremes[well_name] = {}
        
        for model_name in models_for_well:
            # Фильтруем данные для этой модели и параметра давления
            model_pressure_df: pd.DataFrame = df_well[
                (df_well['model'] == model_name) &
                (df_well['parameter'] == 'pressure')
            ].copy()
            
            if model_pressure_df.empty:
                model_extremes[well_name][model_name] = {'maxima': {}, 'minima': {}}
                continue
            
            # Подготавливаем данные для поиска экстремумов (vectorized)
            # Фильтруем NaN значения
            model_pressure_df_clean = model_pressure_df[model_pressure_df['value'].notna()].copy()
            
            if model_pressure_df_clean.empty:
                model_extremes[well_name][model_name] = {'maxima': {}, 'minima': {}}
                continue
            
            # Преобразуем даты в строковый формат (vectorized)
            def format_date(date_val):
                if isinstance(date_val, (pd.Timestamp, datetime)):
                    return date_val.strftime('%d.%m.%Y')
                elif isinstance(date_val, np.datetime64):
                    return pd.Timestamp(date_val).strftime('%d.%m.%Y')
                else:
                    return str(date_val)
            
            model_pressure_df_clean['date_str'] = model_pressure_df_clean['date'].apply(format_date)
            
            pressure_data = [
                {'date': date_str, 'pressure_smoothed': value}
                for date_str, value in zip(model_pressure_df_clean['date_str'], model_pressure_df_clean['value'])
            ]
            
            if len(pressure_data) < 10:
                model_extremes[well_name][model_name] = {'maxima': {}, 'minima': {}}
                continue
            
            # Создаем DataFrame для поиска экстремумов
            pressure_df: pd.DataFrame = pd.DataFrame(pressure_data)
            
            try:
                # Ищем экстремумы с более чувствительными параметрами для модельных данных
                extremes_df: pd.DataFrame = find_extremes_improved_v2(
                    pressure_df,
                    min_distance_days=MODEL_MIN_DISTANCE_DAYS,
                    prominence_percent=MODEL_PROMINENCE_PERCENT,
                    max_cycle_days=MODEL_MAX_CYCLE_DAYS,
                    edge_buffer_days=MODEL_EDGE_BUFFER_DAYS,
                    exclude_end_days=MODEL_EXCLUDE_END_DAYS
                )
                
                # Создаем словари для экстремумов (vectorized)
                def normalize_date_key(date_val):
                    """Нормализует дату в формат YYYY-MM-DD"""
                    try:
                        if isinstance(date_val, str):
                            try:
                                date_dt = datetime.strptime(date_val, '%d.%m.%Y')
                            except Exception:
                                date_dt = pd.to_datetime(date_val)
                        else:
                            date_dt = pd.to_datetime(date_val)
                        return date_dt.strftime('%Y-%m-%d')
                    except Exception:
                        return str(date_val)
                
                # Преобразуем даты в нормализованный формат (vectorized)
                extremes_df['date_key'] = extremes_df['date'].apply(normalize_date_key)
                
                # Фильтруем и создаем словари (vectorized)
                maxima_mask = extremes_df['maxima'].notna()
                minima_mask = extremes_df['minima'].notna()
                
                maxima_dict = dict(zip(
                    extremes_df.loc[maxima_mask, 'date_key'],
                    extremes_df.loc[maxima_mask, 'maxima']
                ))
                
                minima_dict = dict(zip(
                    extremes_df.loc[minima_mask, 'date_key'],
                    extremes_df.loc[minima_mask, 'minima']
                ))
                
                model_extremes[well_name][model_name] = {
                    'maxima': maxima_dict,
                    'minima': minima_dict
                }
                
            except Exception:
                model_extremes[well_name][model_name] = {'maxima': {}, 'minima': {}}
    
    return model_extremes


def collect_last_extremes(
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    num_extremes: int = 5
) -> List[Dict[str, Any]]:
    """
    Собирает последние N максимумов и минимумов для факта и всех моделей.
    
    Parameters
    ----------
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных: {well: {'maxima': {date: value}, 'minima': {date: value}}}
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей: {well: {model: {'maxima': {date: value}, 'minima': {date: value}}}}
    num_extremes : int
        Количество последних экстремумов для сбора (по умолчанию 5)
    
    Returns
    -------
    List[Dict[str, Any]]
        Список словарей с данными: well, model_name, extremum_type, extremum_order, date, wbp
    """
    result: List[Dict[str, Any]] = []
    
    # Получаем все скважины
    all_wells: set = set(extremes_data.keys())
    all_wells.update(model_extremes.keys())
    
    for well_name in all_wells:
        # Обрабатываем фактические данные
        if well_name in extremes_data:
            fact_data = extremes_data[well_name]
            
            # Собираем максимумы факта
            fact_maxima = fact_data.get('maxima', {})
            if fact_maxima:
                # Сортируем по дате (от новых к старым) и берем последние num_extremes
                sorted_maxima = sorted(
                    [(date, value) for date, value in fact_maxima.items()],
                    key=lambda x: x[0],
                    reverse=True
                )[:num_extremes]
                
                # Нумеруем от 1 до num_extremes (1 - самый последний)
                for order, (date_str, value) in enumerate(sorted_maxima, 1):
                    result.append({
                        'well': well_name,
                        'model_name': 'FACT',
                        'extremum_type': 'max',
                        'extremum_order': order,
                        'date': date_str,
                        'wbp': value
                    })
            
            # Собираем минимумы факта
            fact_minima = fact_data.get('minima', {})
            if fact_minima:
                sorted_minima = sorted(
                    [(date, value) for date, value in fact_minima.items()],
                    key=lambda x: x[0],
                    reverse=True
                )[:num_extremes]
                
                for order, (date_str, value) in enumerate(sorted_minima, 1):
                    result.append({
                        'well': well_name,
                        'model_name': 'FACT',
                        'extremum_type': 'min',
                        'extremum_order': order,
                        'date': date_str,
                        'wbp': value
                    })
        
        # Обрабатываем модельные данные
        if well_name in model_extremes:
            for model_name, model_data in model_extremes[well_name].items():
                # Максимумы модели
                model_maxima = model_data.get('maxima', {})
                if model_maxima:
                    sorted_maxima = sorted(
                        [(date, value) for date, value in model_maxima.items()],
                        key=lambda x: x[0],
                        reverse=True
                    )[:num_extremes]
                    
                    for order, (date_str, value) in enumerate(sorted_maxima, 1):
                        result.append({
                            'well': well_name,
                            'model_name': model_name,
                            'extremum_type': 'max',
                            'extremum_order': order,
                            'date': date_str,
                            'wbp': value
                        })
                
                # Минимумы модели
                model_minima = model_data.get('minima', {})
                if model_minima:
                    sorted_minima = sorted(
                        [(date, value) for date, value in model_minima.items()],
                        key=lambda x: x[0],
                        reverse=True
                    )[:num_extremes]
                    
                    for order, (date_str, value) in enumerate(sorted_minima, 1):
                        result.append({
                            'well': well_name,
                            'model_name': model_name,
                            'extremum_type': 'min',
                            'extremum_order': order,
                            'date': date_str,
                            'wbp': value
                        })
    
    return result


def calculate_quality_metrics(
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    num_extremes: int = 5
) -> List[Dict[str, Any]]:
    """
    Вычисляет метрики качества соответствия моделей факту.
    
    Parameters
    ----------
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей
    num_extremes : int
        Количество экстремумов для сравнения (по умолчанию 5)
    
    Returns
    -------
    List[Dict[str, Any]]
        Список словарей с метриками: well, model_name, phase_deviation_days, 
        amplitude_deviation, max_deviation, min_deviation
        Все метрики сохраняют знак для понимания направления отклонения:
        - phase_deviation_days: положительное = модель опережает, отрицательное = отстает
        - amplitude_deviation: положительное = амплитуда модели больше, отрицательное = меньше
        - max_deviation: положительное = максимумы модели выше, отрицательное = ниже
        - min_deviation: положительное = минимумы модели выше, отрицательное = ниже
    """
    result: List[Dict[str, Any]] = []
    
    # Получаем все скважины
    all_wells: set = set(extremes_data.keys())
    all_wells.update(model_extremes.keys())
    
    for well_name in all_wells:
        if well_name not in extremes_data:
            continue
        
        fact_data = extremes_data[well_name]
        fact_maxima = fact_data.get('maxima', {})
        fact_minima = fact_data.get('minima', {})
        
        # Получаем последние N экстремумов факта
        fact_max_sorted = sorted(
            [(date, value) for date, value in fact_maxima.items()],
            key=lambda x: x[0],
            reverse=True
        )[:num_extremes]
        
        fact_min_sorted = sorted(
            [(date, value) for date, value in fact_minima.items()],
            key=lambda x: x[0],
            reverse=True
        )[:num_extremes]
        
        # Обрабатываем каждую модель
        if well_name in model_extremes:
            for model_name, model_data in model_extremes[well_name].items():
                model_maxima = model_data.get('maxima', {})
                model_minima = model_data.get('minima', {})
                
                model_max_sorted = sorted(
                    [(date, value) for date, value in model_maxima.items()],
                    key=lambda x: x[0],
                    reverse=True
                )[:num_extremes]
                
                model_min_sorted = sorted(
                    [(date, value) for date, value in model_minima.items()],
                    key=lambda x: x[0],
                    reverse=True
                )[:num_extremes]
                
                # Вычисляем метрики (без abs, чтобы видеть направление отклонения)
                phase_deviations: List[float] = []
                amplitude_deviations: List[float] = []
                max_deviations: List[float] = []
                min_deviations: List[float] = []
                
                # Сопоставляем максимумы (последний с последним и т.д.)
                num_max_pairs = min(len(fact_max_sorted), len(model_max_sorted))
                for i in range(num_max_pairs):
                    fact_date_str, fact_value = fact_max_sorted[i]
                    model_date_str, model_value = model_max_sorted[i]
                    
                    # Отклонение по фазе (в днях, положительное = модель опережает, отрицательное = отстает)
                    try:
                        fact_date = pd.to_datetime(fact_date_str)
                        model_date = pd.to_datetime(model_date_str)
                        phase_diff = (model_date - fact_date).days  # Без abs для сохранения знака
                        phase_deviations.append(phase_diff)
                    except Exception:
                        pass
                    
                    # Отклонение максимума (положительное = модель выше, отрицательное = ниже)
                    max_deviations.append(model_value - fact_value)
                
                # Сопоставляем минимумы
                num_min_pairs = min(len(fact_min_sorted), len(model_min_sorted))
                for i in range(num_min_pairs):
                    fact_date_str, fact_value = fact_min_sorted[i]
                    model_date_str, model_value = model_min_sorted[i]
                    
                    # Отклонение по фазе
                    try:
                        fact_date = pd.to_datetime(fact_date_str)
                        model_date = pd.to_datetime(model_date_str)
                        phase_diff = (model_date - fact_date).days  # Без abs для сохранения знака
                        phase_deviations.append(phase_diff)
                    except Exception:
                        pass
                    
                    # Отклонение минимума (положительное = модель выше, отрицательное = ниже)
                    min_deviations.append(model_value - fact_value)
                
                # Вычисляем разность амплитуд для пар максимум-минимум
                # Сопоставляем пары экстремумов (последний максимум с последним минимумом и т.д.)
                num_amplitude_pairs = min(num_max_pairs, num_min_pairs)
                for i in range(num_amplitude_pairs):
                    # Амплитуда факта (разница между максимумом и минимумом)
                    fact_max_value = fact_max_sorted[i][1]
                    fact_min_value = fact_min_sorted[i][1]
                    fact_amplitude = fact_max_value - fact_min_value  # Без abs для сохранения знака
                    
                    # Амплитуда модели
                    model_max_value = model_max_sorted[i][1]
                    model_min_value = model_min_sorted[i][1]
                    model_amplitude = model_max_value - model_min_value  # Без abs для сохранения знака
                    
                    # Разность амплитуд (положительное = амплитуда модели больше, отрицательное = меньше)
                    amplitude_diff = model_amplitude - fact_amplitude
                    amplitude_deviations.append(amplitude_diff)
                
                # Вычисляем средние значения
                phase_deviation_days = np.mean(phase_deviations) if phase_deviations else None
                amplitude_deviation = np.mean(amplitude_deviations) if amplitude_deviations else None
                max_deviation = np.mean(max_deviations) if max_deviations else None
                min_deviation = np.mean(min_deviations) if min_deviations else None
                
                result.append({
                    'well': well_name,
                    'model_name': model_name,
                    'phase_deviation_days': phase_deviation_days,
                    'amplitude_deviation': amplitude_deviation,
                    'max_deviation': max_deviation,
                    'min_deviation': min_deviation
                })
    
    return result


def save_to_excel_structured_single_sheet(well_dataframes, historical_df, models_data, 
                                        output_path="structured_comparison_single_sheet.xlsx"):
    """
    Сохранить данные в Excel файл на один лист с указанной структурой в виде единой таблицы:
    1 строка: well, date, wbp_hist, wbp_hist_smoothed, wbp_hist_smoothed_max, wbp_hist_smoothed_min, 
              model_name1, "", "", "", "", model_name2, "", "", "", "", ...
    2 строка: "", "", "", "", "", "", wbp_model, wbp_model_max, wbp_model_min, wgpr_model, wgir_model, 
              wbp_model, wbp_model_max, wbp_model_min, wgpr_model, wgir_model...
    А ниже данные по всем скважинам в единой таблице.
    
    Parameters:
    -----------
    well_dataframes : dict
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные
    models_data : dict
        Сырые данные моделей
    output_path : str
        Путь для сохранения Excel файл
    """
    # Вычисляем сглаженные давления и экстремумы
    smoothed_pressures, extremes_data = compute_smoothed_pressures_and_extremes(well_dataframes, historical_df)
    
    # Вычисляем экстремумы для модельных давлений
    model_extremes = compute_model_extremes(well_dataframes)
    
    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Все скважины"
    
    # Получаем список всех моделей (уникальные по всем скважинам)
    all_models = set()
    for df_well in well_dataframes.values():
        all_models.update([m for m in df_well['model'].unique() if m != 'HISTORICAL'])
    all_models = sorted(all_models)
    
    if not all_models:
        ws.append(["Нет модельных данных"])
        wb.save(output_path)
        return wb, extremes_data, model_extremes
    
    # === СТРОКА 1: Заголовки таблицы ===
    header_row1 = ['well', 'date', 'wbp_hist', 'wbp_hist_smoothed', 'wbp_hist_smoothed_max', 'wbp_hist_smoothed_min']
    for model in all_models:
        # Для каждой модели теперь 5 столбцов
        header_row1.append(model)  # Название модели
        header_row1.append("")      # Пустой столбец для wbp_model_max
        header_row1.append("")      # Пустой столбец для wbp_model_min
        header_row1.append("")      # Пустой столбец для wgpr_model
        header_row1.append("")      # Пустой столбец для wgir_model
    
    ws.append(header_row1)
    
    # === СТРОКА 2: Подзаголовки параметров ===
    header_row2 = ['', '', '', '', '', '']
    for model in all_models:
        header_row2.append('wbp_model')
        header_row2.append('wbp_model_max')
        header_row2.append('wbp_model_min')
        header_row2.append('wgpr_model')
        header_row2.append('wgir_model')
    
    ws.append(header_row2)
    
    # === ПРЕДВАРИТЕЛЬНАЯ ИНДЕКСАЦИЯ ДАННЫХ ДЛЯ БЫСТРОГО ДОСТУПА ===
    # Создаем индексы для быстрого поиска данных по скважинам, датам и моделям
    well_data_indexed: Dict[str, Dict[str, Dict[str, Dict[str, float]]]] = {}
    
    for well_name, df_well in well_dataframes.items():
        well_data_indexed[well_name] = {
            'models': set(df_well['model'].unique()),
            'by_date_model_param': {}
        }
        
        # Индексируем данные по (date, model, parameter) -> value
        for _, row in df_well.iterrows():
            date = row['date']
            model = row['model']
            param = row['parameter']
            value = row['value']
            
            # Нормализуем дату в строку YYYY-MM-DD
            try:
                if isinstance(date, (pd.Timestamp, datetime)):
                    date_key = date.strftime('%Y-%m-%d')
                elif isinstance(date, np.datetime64):
                    date_key = pd.Timestamp(date).strftime('%Y-%m-%d')
                else:
                    date_key = pd.to_datetime(date).strftime('%Y-%m-%d')
            except Exception:
                date_key = str(date)
            
            key = (date_key, model, param)
            if key not in well_data_indexed[well_name]['by_date_model_param']:
                well_data_indexed[well_name]['by_date_model_param'][key] = value
    
    # Нормализуем ключи дат в словарях экстремумов
    def normalize_extremes_dict(extremes_dict: Dict[str, float]) -> Dict[str, float]:
        """Нормализует ключи дат в словаре экстремумов"""
        normalized = {}
        for date_key, value in extremes_dict.items():
            try:
                if isinstance(date_key, str):
                    norm_key = pd.to_datetime(date_key).strftime('%Y-%m-%d')
                else:
                    norm_key = pd.to_datetime(date_key).strftime('%Y-%m-%d')
                normalized[norm_key] = value
            except Exception:
                normalized[str(date_key)] = value
        return normalized
    
    # Нормализуем экстремумы для факта
    normalized_extremes_data: Dict[str, Dict[str, Dict[str, float]]] = {}
    for well_name, well_extremes in extremes_data.items():
        normalized_extremes_data[well_name] = {
            'maxima': normalize_extremes_dict(well_extremes.get('maxima', {})),
            'minima': normalize_extremes_dict(well_extremes.get('minima', {}))
        }
    
    # Нормализуем экстремумы для моделей
    normalized_model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]] = {}
    for well_name, well_models in model_extremes.items():
        normalized_model_extremes[well_name] = {}
        for model_name, model_extremes_dict in well_models.items():
            normalized_model_extremes[well_name][model_name] = {
                'maxima': normalize_extremes_dict(model_extremes_dict.get('maxima', {})),
                'minima': normalize_extremes_dict(model_extremes_dict.get('minima', {}))
            }
    
    # Нормализуем сглаженные давления
    normalized_smoothed_pressures: Dict[str, Dict[str, float]] = {}
    for well_name, well_smoothed in smoothed_pressures.items():
        normalized_smoothed_pressures[well_name] = normalize_extremes_dict(well_smoothed)
    
    # === СБОР ВСЕХ ДАННЫХ В ОДНУ ТАБЛИЦУ (ОПТИМИЗИРОВАННАЯ ВЕРСИЯ) ===
    all_rows: List[List[Any]] = []
    
    for well_name, df_well in well_dataframes.items():
        # Получаем уникальные даты для этой скважины
        well_dates = sorted(df_well['date'].unique())
        
        if not well_dates:
            continue
        
        well_indexed = well_data_indexed[well_name]
        well_models = well_indexed['models']
        
        # Для каждой даты создаем строку данных
        for date in well_dates:
            # Нормализуем дату
            try:
                if isinstance(date, (pd.Timestamp, datetime)):
                    date_key = date.strftime('%Y-%m-%d')
                    date_str = date_key
                elif isinstance(date, np.datetime64):
                    date_ts = pd.Timestamp(date)
                    date_key = date_ts.strftime('%Y-%m-%d')
                    date_str = date_key
                else:
                    date_ts = pd.to_datetime(date)
                    date_key = date_ts.strftime('%Y-%m-%d')
                    date_str = date_key
            except Exception:
                date_key = str(date)
                date_str = date_key
            
            # Получаем историческое давление (быстрый доступ через индекс)
            hist_pressure = well_indexed['by_date_model_param'].get(
                (date_key, 'HISTORICAL', 'pressure'), None
            )
            
            # Получаем сглаженное историческое давление
            hist_smoothed = normalized_smoothed_pressures.get(well_name, {}).get(date_key, None)
            
            # Получаем максимумы и минимумы для сглаженных данных
            well_extremes = normalized_extremes_data.get(well_name, {})
            hist_smoothed_max = well_extremes.get('maxima', {}).get(date_key, None)
            hist_smoothed_min = well_extremes.get('minima', {}).get(date_key, None)
            
            # Создаем строку данных
            data_row = [well_name, date_str, hist_pressure, hist_smoothed, hist_smoothed_max, hist_smoothed_min]
            
            # Добавляем модельные данные для каждой модели
            for model in all_models:
                if model in well_models:
                    # Быстрый доступ через индекс
                    wbp_value = well_indexed['by_date_model_param'].get(
                        (date_key, model, 'pressure'), None
                    )
                    wgpr_value = well_indexed['by_date_model_param'].get(
                        (date_key, model, 'gas_rate'), None
                    )
                    wgir_value = well_indexed['by_date_model_param'].get(
                        (date_key, model, 'gas_injection'), None
                    )
                    
                    # Максимумы и минимумы модельного давления
                    model_extremes_dict = normalized_model_extremes.get(well_name, {}).get(model, {})
                    wbp_model_max = model_extremes_dict.get('maxima', {}).get(date_key, None)
                    wbp_model_min = model_extremes_dict.get('minima', {}).get(date_key, None)
                else:
                    wbp_value = None
                    wbp_model_max = None
                    wbp_model_min = None
                    wgpr_value = None
                    wgir_value = None
                
                data_row.extend([wbp_value, wbp_model_max, wbp_model_min, wgpr_value, wgir_value])
            
            all_rows.append(data_row)
    
    # Добавляем все строки пакетом (значительно быстрее)
    print(f"  Добавление {len(all_rows)} строк данных...")
    for row in all_rows:
        ws.append(row)
    
    total_rows = len(all_rows)
    
    # Сохраняем файл (без форматирования для ускорения)
    try:
        wb.save(output_path)
    except Exception as e:
        print(f"✗ Ошибка при сохранении файла: {e}")
        import traceback
        traceback.print_exc()
    
    return wb, extremes_data, model_extremes


def save_extremes_to_excel_sheet(
    wb: Workbook,
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
) -> None:
    """
    Сохраняет данные о последних экстремумах на второй лист Excel.
    
    Parameters
    ----------
    wb : Workbook
        Рабочая книга Excel
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей
    """
    print("  Сохранение данных об экстремумах на лист 2...")
    
    # Создаем или получаем второй лист
    if len(wb.worksheets) < 2:
        ws = wb.create_sheet("Экстремумы")
    else:
        ws = wb.worksheets[1]
        ws.title = "Экстремумы"
    
    # Заголовки
    ws.append(['well', 'model_name', 'extremum_type', 'extremum_order', 'date', 'wbp'])
    
    # Собираем данные об экстремумах
    extremes_list = collect_last_extremes(extremes_data, model_extremes, num_extremes=5)
    
    # Записываем данные
    for item in extremes_list:
        ws.append([
            item['well'],
            item['model_name'],
            item['extremum_type'],
            item['extremum_order'],
            item['date'],
            item['wbp']
        ])
    
    print(f"    ✓ Сохранено {len(extremes_list)} записей об экстремумах")


def save_quality_metrics_to_excel_sheet(
    wb: Workbook,
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
) -> None:
    """
    Сохраняет метрики качества соответствия моделей факту на третий лист Excel.
    
    Parameters
    ----------
    wb : Workbook
        Рабочая книга Excel
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей
    """
    print("  Сохранение метрик качества на лист 3...")
    
    # Создаем или получаем третий лист
    if len(wb.worksheets) < 3:
        ws = wb.create_sheet("Качество соответствия")
    else:
        ws = wb.worksheets[2]
        ws.title = "Качество соответствия"
    
    # Заголовки
    ws.append([
        'well',
        'model_name',
        'phase_deviation_days',
        'amplitude_deviation',
        'max_deviation',
        'min_deviation'
    ])
    
    # Вычисляем метрики качества
    quality_metrics = calculate_quality_metrics(extremes_data, model_extremes, num_extremes=5)
    
    # Записываем данные
    for item in quality_metrics:
        ws.append([
            item['well'],
            item['model_name'],
            item['phase_deviation_days'],
            item['amplitude_deviation'],
            item['max_deviation'],
            item['min_deviation']
        ])
    
    print(f"    ✓ Сохранено {len(quality_metrics)} записей о качестве соответствия")


def save_to_excel_with_all_sheets(
    well_dataframes: Dict[str, pd.DataFrame],
    historical_df: pd.DataFrame,
    models_data: Dict[str, Dict[str, Any]],
    output_path: str = "structured_comparison_no_interpolation.xlsx"
) -> str:
    """
    Сохраняет данные в Excel файл с тремя листами:
    1. Основные данные (упрощенная версия без форматирования)
    2. Экстремумы (последние 5 максимумов и минимумов)
    3. Качество соответствия (метрики для каждой модели)
    
    Parameters
    ----------
    well_dataframes : Dict[str, pd.DataFrame]
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные
    models_data : Dict[str, Dict[str, Any]]
        Сырые данные моделей
    output_path : str
        Путь для сохранения Excel файла
    
    Returns
    -------
    Tuple[str, Dict[str, Dict[str, Dict[str, float]]], Dict[str, Dict[str, Dict[str, Dict[str, float]]]]]
        (путь к файлу, extremes_data, model_extremes)
    """
    print(f"\nСохранение данных в Excel файл: {output_path}")
    
    # Сохраняем основные данные (первый лист)
    wb, extremes_data, model_extremes = save_to_excel_structured_single_sheet(
        well_dataframes, historical_df, models_data, output_path
    )
    
    # Сохраняем данные об экстремумах (второй лист)
    save_extremes_to_excel_sheet(wb, extremes_data, model_extremes)
    
    # Сохраняем метрики качества (третий лист)
    save_quality_metrics_to_excel_sheet(wb, extremes_data, model_extremes)
    
    # Сохраняем файл с обновленными листами
    try:
        wb.save(output_path)
        print(f"\n✓ Файл успешно сохранен: {output_path}")
        print(f"  Листы: 1. Все скважины, 2. Экстремумы, 3. Качество соответствия")
    except Exception as e:
        print(f"✗ Ошибка при сохранении файла: {e}")
        import traceback
        traceback.print_exc()
    
    return output_path, extremes_data, model_extremes


def prepare_graph_data(
    well_dataframes: Dict[str, pd.DataFrame],
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    num_extremes: int = 5
) -> List[Dict[str, Any]]:
    """
    Подготавливает данные для отправки на сервер построения графиков.
    
    Parameters
    ----------
    well_dataframes : Dict[str, pd.DataFrame]
        DataFrame по скважинам с данными о датах и давлениях
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей
    num_extremes : int
        Количество последних экстремумов для включения (по умолчанию 5)
    
    Returns
    -------
    List[Dict[str, Any]]
        Список словарей с данными для каждой скважины в формате для сервера
    """
    result: List[Dict[str, Any]] = []
    
    # Получаем все скважины
    all_wells: set = set(well_dataframes.keys())
    all_wells.update(extremes_data.keys())
    all_wells.update(model_extremes.keys())
    
    for well_name in all_wells:
        well_data: Dict[str, Any] = {'well_name': well_name}
        
        # Подготавливаем данные для факта
        if well_name in well_dataframes:
            df_well = well_dataframes[well_name]
            
            # Получаем фактические данные (HISTORICAL)
            fact_data_df = df_well[
                (df_well['model'] == 'HISTORICAL') &
                (df_well['parameter'] == 'pressure')
            ].copy()
            
            if not fact_data_df.empty:
                # Сортируем по дате
                fact_data_df = fact_data_df.sort_values('date')
                
                # Извлекаем даты и давления (vectorized)
                # Фильтруем NaN значения
                fact_data_df_clean = fact_data_df[fact_data_df['value'].notna()].copy()
                
                if not fact_data_df_clean.empty:
                    # Преобразуем даты в строки (vectorized)
                    def format_date_to_str(date_val):
                        if isinstance(date_val, (pd.Timestamp, datetime)):
                            return date_val.strftime('%Y-%m-%d')
                        elif isinstance(date_val, np.datetime64):
                            return pd.Timestamp(date_val).strftime('%Y-%m-%d')
                        else:
                            try:
                                return pd.to_datetime(date_val).strftime('%Y-%m-%d')
                            except Exception:
                                return str(date_val)
                    
                    dates = fact_data_df_clean['date'].apply(format_date_to_str).tolist()
                    wbp = fact_data_df_clean['value'].astype(float).tolist()
                else:
                    dates = []
                    wbp = []
                
                well_data['fact'] = {
                    'dates': dates,
                    'wbp': wbp,
                    'extremums': []
                }
                
                # Добавляем экстремумы для факта
                if well_name in extremes_data:
                    fact_extremes = extremes_data[well_name]
                    
                    # Максимумы
                    fact_maxima = fact_extremes.get('maxima', {})
                    if fact_maxima:
                        sorted_maxima = sorted(
                            [(date, value) for date, value in fact_maxima.items()],
                            key=lambda x: x[0],
                            reverse=True
                        )[:num_extremes]
                        
                        for date_str, value in sorted_maxima:
                            well_data['fact']['extremums'].append({
                                'date': date_str,
                                'wbp': value,
                                'type': 'max'
                            })
                    
                    # Минимумы
                    fact_minima = fact_extremes.get('minima', {})
                    if fact_minima:
                        sorted_minima = sorted(
                            [(date, value) for date, value in fact_minima.items()],
                            key=lambda x: x[0],
                            reverse=True
                        )[:num_extremes]
                        
                        for date_str, value in sorted_minima:
                            well_data['fact']['extremums'].append({
                                'date': date_str,
                                'wbp': value,
                                'type': 'min'
                            })
        
        # Подготавливаем данные для моделей
        if well_name in model_extremes and well_name in well_dataframes:
            df_well = well_dataframes[well_name]
            
            for model_name, model_extreme_data in model_extremes[well_name].items():
                # Получаем данные модели
                model_data_df = df_well[
                    (df_well['model'] == model_name) &
                    (df_well['parameter'] == 'pressure')
                ].copy()
                
                if not model_data_df.empty:
                    # Сортируем по дате
                    model_data_df = model_data_df.sort_values('date')
                    
                    # Извлекаем даты и давления (vectorized)
                    # Фильтруем NaN значения
                    model_data_df_clean = model_data_df[model_data_df['value'].notna()].copy()
                    
                    if not model_data_df_clean.empty:
                        # Преобразуем даты в строки (vectorized)
                        def format_date_to_str(date_val):
                            if isinstance(date_val, (pd.Timestamp, datetime)):
                                return date_val.strftime('%Y-%m-%d')
                            elif isinstance(date_val, np.datetime64):
                                return pd.Timestamp(date_val).strftime('%Y-%m-%d')
                            else:
                                try:
                                    return pd.to_datetime(date_val).strftime('%Y-%m-%d')
                                except Exception:
                                    return str(date_val)
                        
                        dates = model_data_df_clean['date'].apply(format_date_to_str).tolist()
                        wbp = model_data_df_clean['value'].astype(float).tolist()
                    else:
                        dates = []
                        wbp = []
                    
                    well_data[model_name] = {
                        'dates': dates,
                        'wbp': wbp,
                        'extremums': []
                    }
                    
                    # Добавляем экстремумы для модели
                    model_maxima = model_extreme_data.get('maxima', {})
                    if model_maxima:
                        sorted_maxima = sorted(
                            [(date, value) for date, value in model_maxima.items()],
                            key=lambda x: x[0],
                            reverse=True
                        )[:num_extremes]
                        
                        for date_str, value in sorted_maxima:
                            well_data[model_name]['extremums'].append({
                                'date': date_str,
                                'wbp': value,
                                'type': 'max'
                            })
                    
                    model_minima = model_extreme_data.get('minima', {})
                    if model_minima:
                        sorted_minima = sorted(
                            [(date, value) for date, value in model_minima.items()],
                            key=lambda x: x[0],
                            reverse=True
                        )[:num_extremes]
                        
                        for date_str, value in sorted_minima:
                            well_data[model_name]['extremums'].append({
                                'date': date_str,
                                'wbp': value,
                                'type': 'min'
                            })
        
        result.append(well_data)
    
    return result


def send_graph_request_and_save_archive(
    well_dataframes: Dict[str, pd.DataFrame],
    extremes_data: Dict[str, Dict[str, Dict[str, float]]],
    model_extremes: Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
) -> Optional[str]:
    """
    Отправляет данные на сервер для построения графиков и сохраняет полученный архив.
    
    Parameters
    ----------
    well_dataframes : Dict[str, pd.DataFrame]
        DataFrame по скважинам
    extremes_data : Dict[str, Dict[str, Dict[str, float]]]
        Экстремумы для фактических данных
    model_extremes : Dict[str, Dict[str, Dict[str, Dict[str, float]]]]
        Экстремумы для моделей
    
    Returns
    -------
    Optional[str]
        Путь к сохраненному архиву или None в случае ошибки
    """
    print(f"\nОтправка данных на сервер для построения графиков...")
    print(f"  Хост: {GRAPH_SERVER_HOST}:{GRAPH_SERVER_PORT}")
    print(f"  Эндпойнт: {GRAPH_SERVER_ENDPOINT}")
    
    try:
        # Подготавливаем данные
        print("  Подготовка данных...")
        graph_data = prepare_graph_data(well_dataframes, extremes_data, model_extremes, num_extremes=5)
        
        if not graph_data:
            print("  Предупреждение: нет данных для отправки")
            return None
        
        print(f"  Подготовлено данных для {len(graph_data)} скважин")
        
        # Формируем JSON
        json_data = json.dumps(graph_data, ensure_ascii=False, indent=2)
        
        # Создаем соединение
        conn = http.client.HTTPConnection(GRAPH_SERVER_HOST, GRAPH_SERVER_PORT)
        
        # Устанавливаем заголовки
        headers = {
            'Content-Type': 'application/json',
            'Content-Length': str(len(json_data.encode('utf-8')))
        }
        
        # Отправляем POST запрос
        print("  Отправка запроса на сервер...")
        conn.request('POST', GRAPH_SERVER_ENDPOINT, json_data, headers)
        
        # Получаем ответ
        response = conn.getresponse()
        
        if response.status != 200:
            print(f"  ✗ Ошибка: сервер вернул статус {response.status}")
            print(f"  Сообщение: {response.read().decode('utf-8')}")
            conn.close()
            return None
        
        # Читаем архив из ответа
        print("  Получение архива с сервера...")
        archive_data = response.read()
        
        conn.close()
        
        # Сохраняем архив
        archive_path = os.path.join(PROJECT_FOLDER_PATH, GRAPH_ARCHIVE_NAME)
        print(f"  Сохранение архива: {archive_path}")
        
        with open(archive_path, 'wb') as f:
            f.write(archive_data)
        
        print(f"  ✓ Архив успешно сохранен: {archive_path}")
        print(f"  Размер архива: {len(archive_data)} байт")
        
        return archive_path
        
    except http.client.HTTPException as e:
        print(f"  ✗ Ошибка HTTP соединения: {e}")
        return None
    except Exception as e:
        print(f"  ✗ Ошибка при отправке запроса: {e}")
        import traceback
        traceback.print_exc()
        return None


def main() -> Optional[Dict[str, Any]]:
    """
    Основная функция для получения унифицированных данных.
    
    Returns
    -------
    Optional[Dict[str, Any]]
        Словарь с результатами обработки или None в случае ошибки
    """
    print("=" * 80)
    print("ПОЛУЧЕНИЕ УНИФИЦИРОВАННЫХ ДАННЫХ ДЛЯ СРАВНЕНИЯ")
    print("=" * 80)
    
    try:
        # 1. Читаем фактические данные из файла
        fact_file_path: str = os.path.join(PROJECT_FOLDER_PATH, WBP_FACT_TXT)
        
        if not os.path.exists(fact_file_path):
            print(f"ОШИБКА: Файл {fact_file_path} не найден!")
            return None
        
        df_fact: pd.DataFrame = parse_fact_well_data(fact_file_path)
        
        if df_fact.empty:
            print("ОШИБКА: Не удалось загрузить фактические данные!")
            return None
        
        print(f"Загружено: {len(df_fact)} записей, {df_fact['well_fact'].nunique()} скважин")
        
        # 2. Получаем унифицированные данные БЕЗ интерполяции
        well_dataframes: Dict[str, pd.DataFrame]
        models_raw: Dict[str, Dict[str, Any]]
        well_dataframes, models_raw = get_unified_data_per_well_without_interpolation(
            model_names=MODEL_NAMES,
            historical_df=df_fact,
            well_column='well_fact',
            date_column='date_fact',
            pressure_column='wpb_bar_fact'
        )
        
        if not well_dataframes:
            print("Ошибка: не удалось получить унифицированные данные!")
            return None
        
        # 3. Вычисляем статистику
        total_records: int = 0
        param_counts: Dict[str, int] = {}
        model_counts: Dict[str, int] = {}
        
        for well, df_well in well_dataframes.items():
            records: int = len(df_well)
            total_records += records
            
            # Считаем параметры
            for param in df_well['parameter'].unique():
                if param not in param_counts:
                    param_counts[param] = 0
                param_counts[param] += len(df_well[df_well['parameter'] == param])
            
            # Считаем модели
            for model in df_well['model'].unique():
                if model not in model_counts:
                    model_counts[model] = 0
                model_counts[model] += len(df_well[df_well['model'] == model])
        
        print(f"\nСтатистика: {total_records} записей, {len(well_dataframes)} скважин")
        
        # 4. Сохраняем данные в Excel с тремя листами
        output_excel: str = os.path.join(
            PROJECT_FOLDER_PATH, "structured_comparison_no_interpolation.xlsx"
        )
        output_excel, extremes_data, model_extremes = save_to_excel_with_all_sheets(
            well_dataframes, df_fact, models_raw, output_excel
        )
        
        # 5. Отправляем данные на сервер для построения графиков
        archive_path = send_graph_request_and_save_archive(
            well_dataframes, extremes_data, model_extremes
        )
        
        # Возвращаем данные для дальнейшего использования
        result = {
            'fact_data': df_fact,
            'well_dataframes': well_dataframes,
            'models_raw': models_raw,
            'excel_file': output_excel,
            'statistics': {
                'total_records': total_records,
                'total_wells': len(well_dataframes),
                'parameter_counts': param_counts,
                'model_counts': model_counts
            }
        }
        
        if archive_path:
            result['graph_archive'] = archive_path
        
        return result
        
    except Exception as e:
        print(f"\nОШИБКА ВО ВРЕМЯ ВЫПОЛНЕНИЯ: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# БЛОК ВЫПОЛНЕНИЯ СКРИПТА
if __name__ == "__main__":
    result_data: Optional[Dict[str, Any]] = main()
    if result_data:
        print("\n" + "=" * 80)
        print("ДАННЫЕ УСПЕШНО ПОДГОТОВЛЕНЫ ДЛЯ СРАВНЕНИЯ!")
        print("=" * 80)
        print(f"Excel файл: {result_data['excel_file']}")