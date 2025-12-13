# Скрипт для анализа качества настройки пластового давления по скважинам
# В качестве входных данных - текстовый файл с фактическими давлениями по формату "Скважина	Дата(ДД.ММ.ГГГГ)	Давления(в барах)"

import os
import warnings
import numpy as np
import pandas as pd
from datetime import datetime
from scipy.signal import savgol_filter
from scipy.interpolate import interp1d
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule


# БЛОК, ЗАПОЛНЯЕМЫЙ ПОЛЬЗОВАТЕЛЕМ
WBP_FACT_TXT: str = "scripts_data/Pfkt0_test.inc"  # Относительный путь до файла с фактическими давлениями
MODEL_NAMES: list[str] = [
    "Hist_L0_adapt_GC_AQ_bz_YAMB_SWL_new_swl1.2",
]

# Глобальные настройки сглаживания
SMOOTHING_WINDOW = 51  # Размер окна сглаживания (нечетное число)
SMOOTHING_POLYORDER = 3  # Порядок полинома для фильтра Савицкого-Голея
SMOOTHING_INTERP_METHOD = 'cubic'  # Метод интерполяции для сглаженной кривой
MIN_POINTS_FOR_SMOOTHING = 5  # Минимальное количество точек для сглаживания
# БЛОК, ЗАПОЛНЯЕМЫЙ ПОЛЬЗОВАТЕЛЕМ

PROJECT_FOLDER_PATH: str = get_project_folder()


def parse_fact_well_data(file_path: str) -> pd.DataFrame:
    """
    Парсит файл с данными скважин и возвращает DataFrame
    """
    print('Формирую датафрейм на основе фактических данных...')
    encodings = ['utf-8', 'cp1251', 'cp1252', 'latin1', 'iso-8859-1', 'windows-1251']

    # Читаем файл
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = f.readlines()
            print(f'Успешно прочитан файл "{file_path}" с кодировкой {encoding}')
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('Файл не читается в представленных кодировках')
    
    data = []
    
    for line in lines:
        # Пропускаем строки с комментариями и служебной информацией
        if line.startswith('--') or line.startswith('PC_') or 'C:\\' in line:
            continue
        
        # Разделяем строку на элементы
        parts = line.strip().split()
        
        # Проверяем, что строка содержит 3 значения (well_fact, date_fact, wpb_bar_fact)
        if len(parts) == 3:
            try:
                well_fact = str(parts[0])
                date_str = parts[1]
                wpb_bar_fact = float(parts[2])
                
                # Преобразуем дату в формат datetime
                try:
                    date_fact = datetime.strptime(date_str, '%d.%m.%Y')
                except ValueError as exc:
                    print(f'ОШИБКА извлечения даты в строке "{line}"')
                    print(f'ОШИБКА: {str(exc)}')
                    continue
                
                data.append({
                    'well_fact': well_fact,
                    'date_fact': date_fact,
                    'wpb_bar_fact': wpb_bar_fact
                })
            except (ValueError, IndexError) as exc:
                print(f'ОШИБКА извлечения данных из строки "{line}"')
                print(f'ОШИБКА: {str(exc)}')
                continue
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Сортируем по well_fact и date_fact для удобства
    if not df.empty:
        df = df.sort_values(['well_fact', 'date_fact']).reset_index(drop=True)
    print('...Сформирован датафрейм на основе фактических данных')
    return df


def get_unique_fact_wells(df) -> list[str]:
    """
    Извлекает список уникальных имен скважин из DataFrame.
    """
    if df is None or df.empty:
        print("DataFrame пустой или None")
        return []
    
    if 'well_fact' not in df.columns:
        print("В DataFrame отсутствует колонка 'well_fact'")
        print(f"Доступные колонки: {list(df.columns)}")
        return []
    
    # Извлекаем уникальные значения и сортируем их
    unique_wells = sorted(df['well_fact'].unique().tolist())
    
    return unique_wells


def get_raw_model_data(model_names: list[str], well_names: list[str], parameters: list[str] | None = None):
    """
    Получить сырые данные из моделей (без интерполяции)
    """
    if parameters is None:
        parameters = ['wbp', 'wgpr', 'wgir']  # Давление, добыча газа, закачка газа
    
    models_data = {}
    
    for model_name in model_names:
        print(f"Загрузка модели: {model_name}")
        
        try:
            # Получаем объект модели
            model = get_model_by_name(model_name)
            
            # Получаем все временные шаги
            timesteps = get_all_timesteps()
            model_dates = [t.to_datetime() for t in timesteps]
            
            # Создаем структуру для данных модели
            model_data = {
                'dates': model_dates,
                'well_data': {}
            }
            
            # Получаем список всех скважин в модели
            try:
                model_wells = get_all_wells()
                model_well_names = [w.name for w in model_wells]
                print(f"  В модели {len(model_well_names)} скважин")
            except:
                print(f"  Не удалось получить список скважин модели")
                model_well_names = []
            
            # Для каждой запрошенной скважины
            for well_name in well_names:
                # Проверяем, есть ли скважина в модели
                if well_name not in model_well_names:
                    print(f"  Предупреждение: скважина {well_name} отсутствует в модели {model_name}")
                    continue
                
                try:
                    well = get_well_by_name(well_name)
                    well_data = {}
                    
                    # Загружаем каждый параметр
                    for param in parameters:
                        try:
                            # Пробуем разные варианты для давления
                            if param == 'wbp':
                                try:
                                    graph_data = wbp[model, well]
                                    print(f"    ✓ {well_name}.{param}: wbp найден")
                                except Exception as e1:
                                    try:
                                        graph_data = wbhp[model, well]
                                        print(f"    ✓ {well_name}.{param}: wbhp найден (вместо wbp)")
                                    except Exception as e2:
                                        try:
                                            graph_data = wbhp_h[model, well]
                                            print(f"    ✓ {well_name}.{param}: wbhp_h найден (вместо wbp)")
                                        except Exception as e3:
                                            print(f"    ✗ {well_name}.{param}: не удалось получить данные давления")
                                            continue
                            elif param == 'wgpr':
                                try:
                                    graph_data = wgpr[model, well]
                                    print(f"    ✓ {well_name}.{param}: найден")
                                except:
                                    print(f"    ✗ {well_name}.{param}: не найден")
                                    continue
                            elif param == 'wgir':
                                try:
                                    graph_data = wgir[model, well]
                                    print(f"    ✓ {well_name}.{param}: найден")
                                except:
                                    print(f"    ✗ {well_name}.{param}: не найден")
                                    continue
                            else:
                                print(f"    ✗ {well_name}.{param}: неподдерживаемый параметр")
                                continue
                            
                            # Извлекаем значения
                            values = []
                            for t in timesteps:
                                try:
                                    value = graph_data[t]
                                    values.append(float(value))
                                except:
                                    values.append(np.nan)
                            
                            well_data[param] = values
                            
                        except Exception as e:
                            print(f"    Ошибка при загрузке параметра {param} для {well_name}: {e}")
                            continue
                    
                    model_data['well_data'][well_name] = well_data
                    print(f"  ✓ Скважина {well_name}: успешно загружена")
                    
                except Exception as e:
                    print(f"  Ошибка при обработке скважины {well_name}: {e}")
                    continue
            
            models_data[model_name] = model_data
            
        except Exception as e:
            print(f"Ошибка при загрузке модели {model_name}: {e}")
            continue
    
    print(f"Итог: загружены данные из {len(models_data)} моделей")
    
    # Выводим сводную статистику по загруженным параметрам
    for model_name, model_data in models_data.items():
        well_count = len(model_data.get('well_data', {}))
        print(f"  {model_name}: {well_count} скважин с данными")
        
        # Статистика по параметрам
        if well_count > 0:
            for well in list(model_data['well_data'].keys())[:3]:  # Первые 3 скважины для примера
                params = model_data['well_data'][well].keys()
                print(f"    {well}: параметры - {list(params)}")
    
    return models_data


def create_combined_dataframe_per_well_without_interpolation(models_raw, historical_df,
                                                           well_column='well_fact',
                                                           date_column='date_fact',
                                                           pressure_column='wpb_bar_fact'):
    """
    Создать объединенный DataFrame с данными, сгруппированными по скважинам
    БЕЗ интерполяции - берем фактические даты и модельные даты как есть
    """
    # Преобразуем исторические даты
    historical_df = historical_df.copy()
    historical_df[date_column] = pd.to_datetime(historical_df[date_column])
    
    well_dataframes = {}
    
    # Сначала получаем список всех скважин
    all_wells = set(historical_df[well_column].unique())
    for model_data in models_raw.values():
        all_wells.update(model_data.get('well_data', {}).keys())
    
    print(f"  Всего уникальных скважин: {len(all_wells)}")
    
    for well in all_wells:
        all_records = []
        
        # Получаем фактические данные для этой скважины
        well_historical = historical_df[historical_df[well_column] == well].copy()
        
        # Определяем минимальную и максимальную даты для фактических данных
        if not well_historical.empty:
            min_fact_date = well_historical[date_column].min()
            max_fact_date = well_historical[date_column].max()
            print(f"    Скважина {well}: фактические даты от {min_fact_date} до {max_fact_date}")
        else:
            min_fact_date = None
            max_fact_date = None
        
        # 1. Добавляем фактические данные для этой скважины
        for _, row in well_historical.iterrows():
            if pd.notna(row.get(pressure_column)):
                all_records.append({
                    'date': row[date_column],
                    'model': 'HISTORICAL',
                    'parameter': 'pressure',
                    'value': row[pressure_column]
                })
        
        # 2. Добавляем модельные данные для этой скважины (без интерполяции)
        for model_name, model_info in models_raw.items():
            if well in model_info.get('well_data', {}):
                model_dates = model_info['dates']
                well_data = model_info['well_data'][well]
                
                # Фильтруем модельные даты по диапазону фактических дат
                if min_fact_date and max_fact_date:
                    # Создаем Series для фильтрации дат
                    date_series = pd.Series(model_dates)
                    date_mask = (date_series >= min_fact_date) & (date_series <= max_fact_date)
                    filtered_dates = date_series[date_mask].tolist()
                    
                    print(f"      Модель {model_name}: отобрано {len(filtered_dates)} из {len(model_dates)} дат")
                else:
                    filtered_dates = model_dates
                
                if not filtered_dates:
                    continue
                
                # Сопоставляем даты с индексами
                date_indices = []
                for date in filtered_dates:
                    try:
                        idx = model_dates.index(date)
                        date_indices.append(idx)
                    except ValueError:
                        continue
                
                # Для каждого параметра
                for param, all_values in well_data.items():
                    if len(all_values) != len(model_dates):
                        print(f"    Предупреждение: несоответствие размеров для модели {model_name}, скважина {well}, параметр {param}")
                        continue
                    
                    # Сопоставляем имена параметров
                    if param == 'wbp':
                        param_display = 'pressure'
                    elif param == 'wgpr':
                        param_display = 'gas_rate'
                    elif param == 'wgir':
                        param_display = 'gas_injection'
                    else:
                        param_display = param
                    
                    # Добавляем записи для отфильтрованных дат
                    for idx in date_indices:
                        date = model_dates[idx]
                        value = all_values[idx]
                        
                        # Пропускаем NaN значения
                        if pd.isna(value):
                            continue
                            
                        all_records.append({
                            'date': date,
                            'model': model_name,
                            'parameter': param_display,
                            'value': value
                        })
        
        # Создаем DataFrame для скважины
        if all_records:
            df_well = pd.DataFrame(all_records)
            df_well = df_well.sort_values(['date', 'parameter', 'model']).reset_index(drop=True)
            
            # Выводим статистику
            fact_count = len(df_well[df_well['model'] == 'HISTORICAL'])
            model_count = len(df_well[df_well['model'] != 'HISTORICAL'])
            print(f"    Скважина {well}: {len(df_well)} записей ({fact_count} факт, {model_count} модель)")
            
            well_dataframes[well] = df_well
        else:
            print(f"    Предупреждение: нет данных для скважины {well}")
    
    return well_dataframes


def get_unified_data_per_well_without_interpolation(model_names, historical_df,
                                                  well_column='well_fact',
                                                  date_column='date_fact',
                                                  pressure_column='wpb_bar_fact'):
    """
    Основная функция: получить унифицированные данные, сгруппированные по скважинам
    БЕЗ интерполяции - берем фактические даты и модельные даты как есть
    """
    print("=" * 60)
    print("ПОЛУЧЕНИЕ УНИФИЦИРОВАННЫХ ДАННЫХ ПО СКВАЖИНАМ (БЕЗ ИНТЕРПОЛЯЦИИ)")
    print("=" * 60)
    
    # 1. Получаем список скважин из фактических данных
    well_names = get_unique_fact_wells(historical_df)
    if not well_names:
        print("Ошибка: не удалось получить список скважин из фактических данных")
        return {}, {}
    
    print(f"Скважины для анализа: {well_names}")
    
    # 2. Загружаем сырые данные из моделей
    print("\n2. Загрузка сырых данных из моделей...")
    models_raw = get_raw_model_data(model_names, well_names, ['wbp', 'wgpr', 'wgir'])
    
    if not models_raw:
        print("Ошибка: не удалось загрузить данные моделей")
        return {}, {}
    
    print(f"Загружены данные из {len(models_raw)} моделей")
    
    # 3. Создаем объединенные DataFrame по скважинам (без интерполяции)
    print("\n3. Создание объединенных DataFrame по скважинам (без интерполяции)...")
    well_dataframes = create_combined_dataframe_per_well_without_interpolation(
        models_raw, historical_df,
        well_column=well_column,
        date_column=date_column,
        pressure_column=pressure_column
    )
    
    print(f"Созданы DataFrame для {len(well_dataframes)} скважин")
    
    # 4. Выводим информацию о загруженных параметрах
    print("\n4. Анализ загруженных параметров...")
    param_stats = {}
    for well, df_well in well_dataframes.items():
        params = df_well['parameter'].unique()
        for param in params:
            if param not in param_stats:
                param_stats[param] = set()
            param_stats[param].add(well)
    
    print("Загружены следующие параметры:")
    for param, wells in param_stats.items():
        print(f"  {param}: {len(wells)} скважин")
    
    return well_dataframes, models_raw


def smooth_pressure_timeseries(df):
    """
    Создает сглаженную кривую для временного ряда давления.
    Сглаживание применяется только к существующим данным, но значения
    сглаженной кривой вычисляются для всех дат.
    
    Параметры:
    -----------
    df : pandas.DataFrame
        Датафрейм с колонками 'date' (строки в формате ДД.ММ.ГГГГ) и 'pressure_fact'
    
    Возвращает:
    -----------
    pandas.DataFrame
        Исходный датафрейм с добавленной колонкой 'pressure_smoothed'
    """
    # Создаем копию датафрейма
    df_result = df.copy()
    
    # Преобразуем даты в datetime
    df_result['date_dt'] = pd.to_datetime(df_result['date'], format='%d.%m.%Y')
    
    # Сортируем по дате
    df_result = df_result.sort_values('date_dt').reset_index(drop=True)
    
    # Шаг 1: Получаем только фактические данные (без NaN)
    fact_mask = df_result['pressure_fact'].notna()
    fact_dates = df_result.loc[fact_mask, 'date_dt']
    fact_pressures = df_result.loc[fact_mask, 'pressure_fact'].values
    
    # Проверяем, достаточно ли точек для сглаживания
    if len(fact_pressures) < MIN_POINTS_FOR_SMOOTHING:
        warnings.warn(f"Недостаточно точек для сглаживания. Доступно: {len(fact_pressures)}, требуется: {MIN_POINTS_FOR_SMOOTHING}")
        # Если точек недостаточно, используем простую линейную интерполяцию
        df_result['pressure_smoothed'] = np.nan
        # Интерполируем только между фактическими точками
        if len(fact_pressures) >= 2:
            # Создаем временную шкалу в днях от первой даты
            all_dates_numeric = (df_result['date_dt'] - df_result['date_dt'].min()).dt.days.values
            fact_dates_numeric = (fact_dates - df_result['date_dt'].min()).dt.days.values
            
            # Линейная интерполяция
            interp_func = interp1d(fact_dates_numeric, fact_pressures, 
                                  kind='linear', bounds_error=False, 
                                  fill_value='extrapolate')
            
            # Получаем значения для всех дат
            smoothed_all = interp_func(all_dates_numeric)
            df_result['pressure_smoothed'] = smoothed_all
        
        df_result.drop(columns=['date_dt'], inplace=True)
        return df_result
    
    # Шаг 2: Создаем числовую временную шкалу для фактических точек
    # Используем дни от первой фактической даты
    first_fact_date = fact_dates.iloc[0]
    fact_dates_numeric = (fact_dates - first_fact_date).dt.days.values
    
    # Шаг 3: Создаем равномерную временную шкалу для сглаживания
    # (фильтр Савицкого-Голея требует равномерной сетки)
    min_date_num = fact_dates_numeric.min()
    max_date_num = fact_dates_numeric.max()
    
    # Создаем равномерную сетку с шагом 1 день
    uniform_dates_num = np.arange(min_date_num, max_date_num + 1)
    
    # Интерполируем фактические данные на равномерную сетку
    # (только для целей сглаживания)
    interp_func_uniform = interp1d(fact_dates_numeric, fact_pressures, 
                                  kind=SMOOTHING_INTERP_METHOD, 
                                  bounds_error=False, 
                                  fill_value='extrapolate')
    
    pressures_uniform = interp_func_uniform(uniform_dates_num)
    
    # Шаг 4: Применяем фильтр Савицкого-Голея к равномерной сетке
    # Настраиваем размер окна
    window_length = min(SMOOTHING_WINDOW, len(pressures_uniform))
    if window_length % 2 == 0:
        window_length -= 1  # Делаем нечетным
        if window_length < 3:
            window_length = 3
    
    # Убедимся, что порядок полинома меньше размера окна
    polyorder = min(SMOOTHING_POLYORDER, window_length - 1)
    
    try:
        smoothed_uniform = savgol_filter(pressures_uniform, 
                                        window_length=window_length,
                                        polyorder=polyorder,
                                        mode='interp')
    except Exception as e:
        warnings.warn(f"Ошибка при сглаживании: {str(e)}. Использую интерполированные значения.")
        smoothed_uniform = pressures_uniform
    
    # Шаг 5: Создаем интерполяционную функцию для сглаженной кривой
    # (снова используем фактическую временную шкалу)
    # Создаем функцию, которая интерполирует сглаженные значения
    # на основе равномерной сетки
    interp_smoothed = interp1d(uniform_dates_num, smoothed_uniform,
                              kind=SMOOTHING_INTERP_METHOD,
                              bounds_error=False,
                              fill_value='extrapolate')
    
    # Шаг 6: Вычисляем сглаженные значения для всех дат (включая пропуски)
    all_dates_numeric = (df_result['date_dt'] - first_fact_date).dt.days.values
    smoothed_all = interp_smoothed(all_dates_numeric)
    
    # Ограничиваем значения разумными пределами
    min_pressure = max(0, np.nanmin(fact_pressures) * 0.5)
    max_pressure = np.nanmax(fact_pressures) * 1.5
    smoothed_all = np.clip(smoothed_all, min_pressure, max_pressure)
    
    df_result['pressure_smoothed'] = smoothed_all
    
    # Шаг 7: Убедимся, что сглаженная кривая не имеет NaN
    # (интерполяция с fill_value='extrapolate' должна это обеспечить)
    if df_result['pressure_smoothed'].isna().any():
        # Заполняем оставшиеся NaN линейной интерполяцией
        df_result['pressure_smoothed'] = df_result['pressure_smoothed'].interpolate(
            method='linear', limit_direction='both'
        )
    
    # Удаляем вспомогательные колонки
    df_result.drop(columns=['date_dt'], inplace=True)
    
    return df_result


def compute_smoothed_pressures(well_dataframes, historical_df):
    """
    Вычисляет сглаженные значения давления для каждой скважины
    
    Parameters:
    -----------
    well_dataframes : dict
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные с фактическими давлениями
        
    Returns:
    --------
    dict
        Словарь с ключами - имена скважин, значения - словари с датами и сглаженными значениями
    """
    smoothed_pressures = {}
    
    for well_name, df_well in well_dataframes.items():
        # Извлекаем исторические данные для этой скважины
        historical_data_well = historical_df[historical_df['well_fact'] == well_name].copy()
        
        if historical_data_well.empty:
            print(f"    Предупреждение: нет исторических данных для скважины {well_name}")
            smoothed_pressures[well_name] = {}
            continue
        
        # Получаем все уникальные даты для этой скважины из well_dataframes
        # (и исторические, и модельные)
        all_dates_for_well = df_well['date'].unique()
        
        if len(all_dates_for_well) == 0:
            print(f"    Предупреждение: для скважины {well_name} нет дат в well_dataframes")
            smoothed_pressures[well_name] = {}
            continue
        
        # Создаем DataFrame с ВСЕМИ датами (не только историческими)
        # в формате, ожидаемом функцией smooth_pressure_timeseries
        pressure_data = []
        
        # Преобразуем исторические даты в datetime для корректного сравнения
        historical_data_well['date_fact_dt'] = pd.to_datetime(historical_data_well['date_fact'])
        
        for date in all_dates_for_well:
            # Приводим все даты к единому формату datetime
            if isinstance(date, (pd.Timestamp, datetime)):
                date_dt = date
            elif isinstance(date, np.datetime64):
                date_dt = pd.Timestamp(date)
            else:
                try:
                    date_dt = pd.to_datetime(date)
                except:
                    print(f"    Предупреждение: не удалось преобразовать дату {date} для скважины {well_name}")
                    continue
            
            # Ищем историческое давление для этой даты
            hist_pressure = None
            
            # Ищем точное совпадение даты
            exact_match = historical_data_well[historical_data_well['date_fact_dt'] == date_dt]
            if not exact_match.empty:
                hist_pressure = exact_match.iloc[0]['wpb_bar_fact']
            else:
                # Ищем ближайшую дату (в пределах 1 дня)
                time_diffs = (historical_data_well['date_fact_dt'] - date_dt).abs()
                min_diff_idx = time_diffs.idxmin() if not time_diffs.empty else None
                if min_diff_idx is not None and time_diffs[min_diff_idx] <= pd.Timedelta(days=1):
                    hist_pressure = historical_data_well.loc[min_diff_idx, 'wpb_bar_fact']
            
            pressure_data.append({
                'date': date_dt.strftime('%d.%m.%Y'),
                'pressure_fact': hist_pressure
            })
        
        if not pressure_data:
            print(f"    Предупреждение: не удалось создать данные для сглаживания скважины {well_name}")
            smoothed_pressures[well_name] = {}
            continue
        
        # Создаем DataFrame для сглаживания
        pressure_df = pd.DataFrame(pressure_data)
        
        # Применяем сглаживание
        try:
            smoothed_df = smooth_pressure_timeseries(pressure_df)
            
            # Создаем словарь для быстрого доступа к сглаженным значениям по дате
            smoothed_dict = {}
            for _, row in smoothed_df.iterrows():
                date_str = row['date']
                try:
                    # Преобразуем строку даты обратно в datetime
                    date_dt = datetime.strptime(date_str, '%d.%m.%Y')
                    smoothed_value = row['pressure_smoothed']
                    
                    # Приводим к тому же типу даты, что и в well_dataframes
                    if isinstance(all_dates_for_well[0], (pd.Timestamp, datetime)):
                        smoothed_dict[date_dt] = smoothed_value
                    elif isinstance(all_dates_for_well[0], np.datetime64):
                        smoothed_dict[np.datetime64(date_dt)] = smoothed_value
                    else:
                        smoothed_dict[date_dt] = smoothed_value
                        
                except Exception as e:
                    print(f"    Ошибка преобразования даты {date_str} для скважины {well_name}: {e}")
                    continue
            
            smoothed_pressures[well_name] = smoothed_dict
            print(f"    ✓ Скважина {well_name}: вычислено {len(smoothed_dict)} сглаженных значений")
            
            # Отладочная информация для первых нескольких значений
            if len(smoothed_dict) > 0:
                sample_dates = list(smoothed_dict.keys())[:3]
                print(f"      Пример: {sample_dates[0].strftime('%d.%m.%Y') if isinstance(sample_dates[0], datetime) else sample_dates[0]} -> {smoothed_dict[sample_dates[0]]:.2f}")
            
        except Exception as e:
            print(f"    Ошибка при сглаживании данных для скважины {well_name}: {str(e)}")
            import traceback
            traceback.print_exc()
            smoothed_pressures[well_name] = {}
    
    return smoothed_pressures


def save_to_excel_structured_single_sheet(well_dataframes, historical_df, models_data, 
                                        output_path="structured_comparison_single_sheet.xlsx"):
    """
    Сохранить данные в Excel файл на один лист с указанной структурой в виде единой таблицы:
    1 строка: well, date, wbp_hist, wbp_hist_smoothed, model_name1, "", "", model_name2, "", "", ...
    2 строка: "", "", "", "", wbp_model, wgpr_model, wgir_model, wbp_model, wgpr_model, wgir_model...
    А ниже данные по всем скважинам в единой таблице.
    
    Parameters:
    -----------
    well_dataframes : dict
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные
    models_data : dict
        Сырые данные моделей
    output_path : str
        Путь для сохранения Excel файла
    """
    print(f"\nСохранение данных в Excel файл (единая таблица): {output_path}")
    
    # Вычисляем сглаженные давления
    print("  Вычисление сглаженных давлений...")
    smoothed_pressures = compute_smoothed_pressures(well_dataframes, historical_df)
    
    # Отладочная информация о сглаженных данных
    print(f"  Информация о сглаженных данных:")
    for well_name, smoothed_dict in smoothed_pressures.items():
        if smoothed_dict:
            print(f"    {well_name}: {len(smoothed_dict)} значений")
        else:
            print(f"    {well_name}: НЕТ сглаженных данных")
    
    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Все скважины"
    
    # Получаем список всех моделей (уникальные по всем скважинам)
    all_models = set()
    for df_well in well_dataframes.values():
        all_models.update([m for m in df_well['model'].unique() if m != 'HISTORICAL'])
    all_models = sorted(all_models)
    
    if not all_models:
        print("    Предупреждение: нет модельных данных")
        ws.append(["Нет модельных данных"])
        wb.save(output_path)
        return output_path
    
    print(f"  Найдено моделей: {all_models}")
    print(f"  Всего скважин: {len(well_dataframes)}")
    
    # === СТРОКА 1: Заголовки таблицы ===
    header_row1 = ['well', 'date', 'wbp_hist', 'wbp_hist_smoothed']
    for model in all_models:
        header_row1.append(model)  # Название модели
        header_row1.append("")      # Пустой столбец
        header_row1.append("")      # Пустой столбец
    
    ws.append(header_row1)
    
    # === СТРОКА 2: Подзаголовки параметров ===
    header_row2 = ['', '', '', '']
    for model in all_models:
        header_row2.append('wbp_model')
        header_row2.append('wgpr_model')
        header_row2.append('wgir_model')
    
    ws.append(header_row2)
    
    # === СБОР ВСЕХ ДАННЫХ В ОДНУ ТАБЛИЦУ ===
    total_rows = 0
    
    for well_idx, (well_name, df_well) in enumerate(well_dataframes.items()):
        print(f"  Обработка скважины {well_name} ({well_idx+1}/{len(well_dataframes)})...")
        
        # Получаем уникальные даты для этой скважины (уже отсортированы)
        well_dates = sorted(df_well['date'].unique())
        
        if not well_dates:
            print(f"    Предупреждение: для скважины {well_name} нет данных по датам")
            continue
        
        # Для каждой даты создаем строку данных
        for date in well_dates:
            # Преобразуем дату в строку (обрабатываем разные форматы дат)
            try:
                if isinstance(date, (pd.Timestamp, datetime)):
                    date_str = date.strftime('%d.%m.%Y')
                    date_key = date  # Используем как есть для поиска
                elif isinstance(date, np.datetime64):
                    # Преобразуем numpy.datetime64 в pandas.Timestamp
                    date_ts = pd.Timestamp(date)
                    date_str = date_ts.strftime('%d.%m.%Y')
                    date_key = date_ts  # Используем Timestamp для поиска
                else:
                    # Пробуем преобразовать строку в datetime
                    date_dt = pd.to_datetime(date)
                    date_str = date_dt.strftime('%d.%m.%Y')
                    date_key = date_dt
            except Exception as e:
                print(f"    Ошибка преобразования даты {date}: {e}")
                date_str = str(date)
                date_key = date
            
            # Получаем историческое давление для этой даты
            hist_pressure = None
            hist_row = df_well[(df_well['date'] == date) & 
                              (df_well['model'] == 'HISTORICAL') & 
                              (df_well['parameter'] == 'pressure')]
            if not hist_row.empty:
                hist_pressure = hist_row['value'].iloc[0]
            
            # Получаем сглаженное историческое давление для этой даты
            hist_smoothed = None
            if well_name in smoothed_pressures:
                # Пробуем найти точное совпадение даты
                if date_key in smoothed_pressures[well_name]:
                    hist_smoothed = smoothed_pressures[well_name][date_key]
                else:
                    # Пробуем найти совпадение с преобразованной датой
                    for dict_date in smoothed_pressures[well_name].keys():
                        try:
                            if isinstance(dict_date, (pd.Timestamp, datetime)):
                                dict_date_str = dict_date.strftime('%d.%m.%Y')
                            elif isinstance(dict_date, np.datetime64):
                                dict_date_str = pd.Timestamp(dict_date).strftime('%d.%m.%Y')
                            else:
                                dict_date_str = str(dict_date)
                            
                            if dict_date_str == date_str:
                                hist_smoothed = smoothed_pressures[well_name][dict_date]
                                break
                        except:
                            continue
            
            # Создаем строку данных
            data_row = [well_name, date_str, hist_pressure, hist_smoothed]
            
            # Добавляем модельные данные для каждой модели из общего списка
            for model in all_models:
                # Проверяем, есть ли данные для этой модели у текущей скважины
                if model in df_well['model'].unique():
                    # Давление (wbp)
                    wbp_value = None
                    wbp_row = df_well[(df_well['date'] == date) & 
                                     (df_well['model'] == model) & 
                                     (df_well['parameter'] == 'pressure')]
                    if not wbp_row.empty:
                        wbp_value = wbp_row['value'].iloc[0]
                    
                    # Добыча газа (wgpr)
                    wgpr_value = None
                    wgpr_row = df_well[(df_well['date'] == date) & 
                                      (df_well['model'] == model) & 
                                      (df_well['parameter'] == 'gas_rate')]
                    if not wgpr_row.empty:
                        wgpr_value = wgpr_row['value'].iloc[0]
                    
                    # Закачка газа (wgir)
                    wgir_value = None
                    wgir_row = df_well[(df_well['date'] == date) & 
                                      (df_well['model'] == model) & 
                                      (df_well['parameter'] == 'gas_injection')]
                    if not wgir_row.empty:
                        wgir_value = wgir_row['value'].iloc[0]
                else:
                    # Если модель отсутствует для этой скважины, заполняем None
                    wbp_value = None
                    wgpr_value = None
                    wgir_value = None
                
                data_row.extend([wbp_value, wgpr_value, wgir_value])
            
            # Добавляем строку в таблицу
            ws.append(data_row)
            total_rows += 1
        
        print(f"    ✓ Добавлено {len(well_dates)} строк для скважины {well_name}")
    
    # Автонастройка ширины столбцов
    print("  Настройка ширины столбцов...")
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Определяем последнюю колонку для фильтра
    last_col = openpyxl.utils.get_column_letter(ws.max_column)
    
    # Добавляем фильтры и закрепляем заголовки
    print("  Применение фильтров...")
    ws.auto_filter.ref = f"A1:{last_col}2"  # Фильтр на первых двух строках заголовков
    ws.freeze_panes = "A3"  # Закрепляем первые две строки заголовков
    
    # Добавляем информационную строку в начало
    ws.insert_rows(1)
    
    info_cell = ws['A1']
    info_cell.value = f"Сравнение моделей и фактических данных | Скважин: {len(well_dataframes)} | Строк данных: {total_rows} | Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    info_cell.font = Font(bold=True, color="FFFFFF")
    info_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    info_cell.alignment = Alignment(horizontal="center")
    
    # Объединяем ячейки информационной строки
    ws.merge_cells(f'A1:{last_col}1')
    
    # Форматирование заголовков таблицы (строки 2 и 3)
    # Строка 2: Основные заголовки
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Строка 3: Подзаголовки параметров
    for cell in ws[3]:
        if cell.column > 4:  # Начиная с 5-го столбца (после well, date, wbp_hist, wbp_hist_smoothed)
            cell.font = Font(italic=True, bold=True)
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Добавляем условное форматирование для столбца сглаженных данных (столбец D)
    # Зеленый цвет для значений
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Применяем условное форматирование ко всем ячейкам в столбце D (начиная с 4 строки)
    for row in range(4, ws.max_row + 1):
        cell = ws[f'D{row}']
        if cell.value is not None:
            cell.fill = green_fill
    
    # Сохраняем файл
    try:
        wb.save(output_path)
        print(f"\n✓ Файл успешно сохранен: {output_path}")
        print(f"  Структура таблицы:")
        print(f"    - Лист: '{ws.title}'")
        print(f"    - Скважин: {len(well_dataframes)}")
        print(f"    - Моделей: {len(all_models)}")
        print(f"    - Столбцов: {ws.max_column}")
        print(f"    - Всего строк данных: {total_rows}")
        print(f"    - Фильтры: строки 2-3")
        print(f"    - Закреплено: строки 1-3")
        
        # Выводим структуру столбцов
        print(f"\n  Структура столбцов:")
        print(f"    1. well - имя скважины")
        print(f"    2. date - дата")
        print(f"    3. wbp_hist - историческое давление")
        print(f"    4. wbp_hist_smoothed - сглаженное историческое давление (зеленый фон)")
        col_idx = 5
        for i, model in enumerate(all_models):
            print(f"    {col_idx}. {model}_wbp - давление модели {model}")
            print(f"    {col_idx+1}. {model}_wgpr - добыча газа модели {model}")
            print(f"    {col_idx+2}. {model}_wgir - закачка газа модели {model}")
            col_idx += 3
        
        # Проверяем, есть ли сглаженные данные в файле
        print(f"\n  Проверка сглаженных данных в файле:")
        smoothed_values_count = 0
        for row in range(4, min(ws.max_row + 1, 20)):  # Проверяем первые 20 строк данных
            cell = ws[f'D{row}']
            if cell.value is not None:
                smoothed_values_count += 1
                if smoothed_values_count <= 5:  # Выводим первые 5 значений
                    well_name = ws[f'A{row}'].value
                    date_val = ws[f'B{row}'].value
                    print(f"    Строка {row}: {well_name}, {date_val} -> {cell.value:.2f}")
        
        print(f"    Всего строк со сглаженными данными в первых {min(ws.max_row - 3, 20)} строках: {smoothed_values_count}")
        
    except Exception as e:
        print(f"✗ Ошибка при сохранении файла: {e}")
        import traceback
        traceback.print_exc()
    
    return output_path

def main():
    """
    Основная функция для получения унифицированных данных
    """
    print("=" * 80)
    print("ПОЛУЧЕНИЕ УНИФИЦИРОВАННЫХ ДАННЫХ ДЛЯ СРАВНЕНИЯ")
    print("=" * 80)
    
    try:
        # 1. Читаем фактические данные из файла
        print(f"\n1. Чтение фактических данных из файла: {WBP_FACT_TXT}")
        fact_file_path = os.path.join(PROJECT_FOLDER_PATH, WBP_FACT_TXT)
        
        if not os.path.exists(fact_file_path):
            print(f"ОШИБКА: Файл {fact_file_path} не найден!")
            print(f"Текущая директория проекта: {PROJECT_FOLDER_PATH}")
            return None
        
        df_fact = parse_fact_well_data(fact_file_path)
        
        if df_fact.empty:
            print("ОШИБКА: Не удалось загрузить фактические данные!")
            return None
        
        print(f"Загружено {len(df_fact)} записей фактических данных")
        print(f"Уникальных скважин: {df_fact['well_fact'].nunique()}")
        
        # 2. Получаем унифицированные данные БЕЗ интерполяции
        print(f"\n2. Получение унифицированных данных (без интерполяции)...")
        well_dataframes, models_raw = get_unified_data_per_well_without_interpolation(
            model_names=MODEL_NAMES,
            historical_df=df_fact,
            well_column='well_fact',
            date_column='date_fact',
            pressure_column='wpb_bar_fact'
        )
        
        if not well_dataframes:
            print("ОШИБКА: Не удалось получить унифицированные данные!")
            return None
        
        # 3. Выводим статистику
        print("\n" + "=" * 60)
        print("РЕЗУЛЬТАТЫ")
        print("=" * 60)
        
        total_records = 0
        param_counts = {}
        model_counts = {}
        
        for well, df_well in well_dataframes.items():
            records = len(df_well)
            total_records += records
            
            # Считаем параметры
            for param in df_well['parameter'].unique():
                if param not in param_counts:
                    param_counts[param] = 0
                param_counts[param] += len(df_well[df_well['parameter'] == param])
            
            # Считаем модели
            for model in df_well['model'].unique():
                if model not in model_counts:
                    model_counts[model] = 0
                model_counts[model] += len(df_well[df_well['model'] == model])
            
            fact_dates = df_well[df_well['model'] == 'HISTORICAL']['date'].nunique()
            model_dates = df_well[df_well['model'] != 'HISTORICAL']['date'].nunique()
            
            print(f"Скважина {well}: {records} записей, "
                  f"{df_well['parameter'].nunique()} параметров, "
                  f"дат: {fact_dates} факт + {model_dates} модель")
        
        print(f"\nВсего записей: {total_records}")
        print(f"Всего скважин: {len(well_dataframes)}")
        print(f"Распределение по параметрам:")
        for param, count in param_counts.items():
            print(f"  {param}: {count} записей")
        
        print(f"\nРаспределение по источникам данных:")
        for model, count in model_counts.items():
            print(f"  {model}: {count} записей")
        
        # 4. Сохраняем данные в Excel с указанной структурой
        print(f"\n3. Сохранение данных в Excel файл...")
        output_excel = os.path.join(PROJECT_FOLDER_PATH, "structured_comparison_no_interpolation.xlsx")
        save_to_excel_structured_single_sheet(well_dataframes, df_fact, models_raw, output_excel)
        
        # Возвращаем данные для дальнейшего использования
        return {
            'fact_data': df_fact,
            'well_dataframes': well_dataframes,
            'models_raw': models_raw,
            'excel_file': output_excel,
            'statistics': {
                'total_records': total_records,
                'total_wells': len(well_dataframes),
                'parameter_counts': param_counts,
                'model_counts': model_counts
            }
        }
        
    except Exception as e:
        print(f"\nОШИБКА ВО ВРЕМЯ ВЫПОЛНЕНИЯ: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# БЛОК ВЫПОЛНЕНИЯ СКРИПТА
result_data = main()
if result_data:
    print("\n" + "=" * 80)
    print("ДАННЫЕ УСПЕШНО ПОДГОТОВЛЕНЫ ДЛЯ СРАВНЕНИЯ!")
    print("=" * 80)
    print(f"Структура данных содержит:")
    print(f"  - Фактические данные: {len(result_data['fact_data'])} записей")
    print(f"  - Данные по скважинам: {len(result_data['well_dataframes'])} скважин")
    print(f"  - Сырые модельные данные: {len(result_data['models_raw'])} моделей")
    print(f"  - Всего записей: {result_data['statistics']['total_records']}")
    print(f"  - Параметры: {list(result_data['statistics']['parameter_counts'].keys())}")
    print(f"  - Excel файл: {result_data['excel_file']}")
    
    # Пример доступа к данным
    if result_data['well_dataframes']:
        first_well = list(result_data['well_dataframes'].keys())[0]
        df_first_well = result_data['well_dataframes'][first_well]
        print(f"\nПример данных для скважины {first_well}:")
        print(f"  Всего записей: {len(df_first_well)}")
        print(f"  Уникальные даты: {df_first_well['date'].nunique()}")
        print(f"  Модели: {df_first_well['model'].unique().tolist()}")
        print(f"  Параметры: {df_first_well['parameter'].unique().tolist()}")
        
        # Выводим первые 10 строк
        print(f"\nПервые 10 строк данных:")
        print(df_first_well.head(10))