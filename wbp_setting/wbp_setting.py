#Automaticaly recalculate=true
#Single model=false
#Run for one model=false
# Скрипт для анализа качества настройки пластового давления по скважинам
# В качестве входных данных - текстовый файл с фактическими давлениями по формату "Скважина	Дата(ДД.ММ.ГГГГ)	Давления(в барах)"

import os
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# БЛОК, ЗАПОЛНЯЕМЫЙ ПОЛЬЗОВАТЕЛЕМ
WBP_FACT_TXT: str = "scripts_data/Pfkt0.inc"  # Относительный путь до файла с фактическими давлениями
MODEL_NAMES: list[str] = [
    "Hist_L0_adapt_GC_AQ_bz_YAMB_SWL_new_swl1.2",
]
# БЛОК, ЗАПОЛНЯЕМЫЙ ПОЛЬЗОВАТЕЛЕМ

PROJECT_FOLDER_PATH: str = get_project_folder()


def parse_fact_well_data(file_path: str) -> pd.DataFrame:
    """
    Парсит файл с данными скважин и возвращает DataFrame
    """
    print('Формирую датафрейм на основе фактических данных...')
    encodings = ['utf-8', 'cp1251', 'cp1252', 'latin1', 'iso-8859-1', 'windows-1251']

    # Читаем файл
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                lines = f.readlines()
            print(f'Успешно прочитан файл "{file_path}" с кодировкой {encoding}')
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('Файл не читается в представленных кодировках')
    
    data = []
    
    for line in lines:
        # Пропускаем строки с комментариями и служебной информацией
        if line.startswith('--') or line.startswith('PC_') or 'C:\\' in line:
            continue
        
        # Разделяем строку на элементы
        parts = line.strip().split()
        
        # Проверяем, что строка содержит 3 значения (well_fact, date_fact, wpb_bar_fact)
        if len(parts) == 3:
            try:
                well_fact = str(parts[0])
                date_str = parts[1]
                wpb_bar_fact = float(parts[2])
                
                # Преобразуем дату в формат datetime
                try:
                    date_fact = datetime.strptime(date_str, '%d.%m.%Y')
                except ValueError as exc:
                    print(f'ОШИБКА извлечения даты в строке "{line}"')
                    print(f'ОШИБКА: {str(exc)}')
                    continue
                
                data.append({
                    'well_fact': well_fact,
                    'date_fact': date_fact,
                    'wpb_bar_fact': wpb_bar_fact
                })
            except (ValueError, IndexError) as exc:
                print(f'ОШИБКА извлечения данных из строки "{line}"')
                print(f'ОШИБКА: {str(exc)}')
                continue
    
    # Создаем DataFrame
    df = pd.DataFrame(data)
    
    # Сортируем по well_fact и date_fact для удобства
    if not df.empty:
        df = df.sort_values(['well_fact', 'date_fact']).reset_index(drop=True)
    print('...Сформирован датафрейм на основе фактических данных')
    return df


def get_unique_fact_wells(df) -> list[str]:
    """
    Извлекает список уникальных имен скважин из DataFrame.
    """
    if df is None or df.empty:
        print("DataFrame пустой или None")
        return []
    
    if 'well_fact' not in df.columns:
        print("В DataFrame отсутствует колонка 'well_fact'")
        print(f"Доступные колонки: {list(df.columns)}")
        return []
    
    # Извлекаем уникальные значения и сортируем их
    unique_wells = sorted(df['well_fact'].unique().tolist())
    
    return unique_wells


def get_well_historical_dates(df_historical, well_column='well_fact', date_column='date_fact') -> dict:
    """
    Получить уникальные даты для каждой скважины из исторических данных
    """
    if df_historical.empty:
        print("Ошибка: DataFrame с историческими данными пустой")
        return {}
    
    # Преобразуем даты в datetime
    df_historical = df_historical.copy()
    df_historical[date_column] = pd.to_datetime(df_historical[date_column])
    
    # Группируем по скважинам и получаем уникальные даты
    well_dates = {}
    
    for well in df_historical[well_column].unique():
        well_data = df_historical[df_historical[well_column] == well]
        dates = sorted(well_data[date_column].dropna().unique())
        well_dates[well] = dates
    
    print(f"Получены даты для {len(well_dates)} скважин")
    
    return well_dates


def get_raw_model_data(model_names: list[str], well_names: list[str], parameters: list[str] | None = None):
    """
    Получить сырые данные из моделей (без интерполяции)
    """
    if parameters is None:
        parameters = ['wbp', 'wgpr', 'wgir']  # Давление, добыча газа, закачка газа
    
    models_data = {}
    
    for model_name in model_names:
        print(f"Загрузка модели: {model_name}")
        
        try:
            # Получаем объект модели
            model = get_model_by_name(model_name)
            
            # Получаем все временные шаги
            timesteps = get_all_timesteps()
            model_dates = [t.to_datetime() for t in timesteps]
            
            # Создаем структуру для данных модели
            model_data = {
                'dates': model_dates,
                'well_data': {}
            }
            
            # Получаем список всех скважин в модели
            try:
                model_wells = get_all_wells()
                model_well_names = [w.name for w in model_wells]
                print(f"  В модели {len(model_well_names)} скважин")
            except:
                print(f"  Не удалось получить список скважин модели")
                model_well_names = []
            
            # Для каждой запрошенной скважины
            for well_name in well_names:
                # Проверяем, есть ли скважина в модели
                if well_name not in model_well_names:
                    print(f"  Предупреждение: скважина {well_name} отсутствует в модели {model_name}")
                    continue
                
                try:
                    well = get_well_by_name(well_name)
                    well_data = {}
                    
                    # Загружаем каждый параметр
                    for param in parameters:
                        try:
                            # Пробуем разные варианты для давления
                            if param == 'wbp':
                                try:
                                    graph_data = wbp[model, well]
                                    print(f"    ✓ {well_name}.{param}: wbp найден")
                                except Exception as e1:
                                    try:
                                        graph_data = wbhp[model, well]
                                        print(f"    ✓ {well_name}.{param}: wbhp найден (вместо wbp)")
                                    except Exception as e2:
                                        try:
                                            graph_data = wbhp_h[model, well]
                                            print(f"    ✓ {well_name}.{param}: wbhp_h найден (вместо wbp)")
                                        except Exception as e3:
                                            print(f"    ✗ {well_name}.{param}: не удалось получить данные давления")
                                            continue
                            elif param == 'wgpr':
                                try:
                                    graph_data = wgpr[model, well]
                                    print(f"    ✓ {well_name}.{param}: найден")
                                except:
                                    print(f"    ✗ {well_name}.{param}: не найден")
                                    continue
                            elif param == 'wgir':
                                try:
                                    graph_data = wgir[model, well]
                                    print(f"    ✓ {well_name}.{param}: найден")
                                except:
                                    print(f"    ✗ {well_name}.{param}: не найден")
                                    continue
                            else:
                                print(f"    ✗ {well_name}.{param}: неподдерживаемый параметр")
                                continue
                            
                            # Извлекаем значения
                            values = []
                            for t in timesteps:
                                try:
                                    value = graph_data[t]
                                    values.append(float(value))
                                except:
                                    values.append(np.nan)
                            
                            well_data[param] = values
                            
                        except Exception as e:
                            print(f"    Ошибка при загрузке параметра {param} для {well_name}: {e}")
                            continue
                    
                    model_data['well_data'][well_name] = well_data
                    print(f"  ✓ Скважина {well_name}: успешно загружена")
                    
                except Exception as e:
                    print(f"  Ошибка при обработке скважины {well_name}: {e}")
                    continue
            
            models_data[model_name] = model_data
            
        except Exception as e:
            print(f"Ошибка при загрузке модели {model_name}: {e}")
            continue
    
    print(f"Итог: загружены данные из {len(models_data)} моделей")
    
    # Выводим сводную статистику по загруженным параметрам
    for model_name, model_data in models_data.items():
        well_count = len(model_data.get('well_data', {}))
        print(f"  {model_name}: {well_count} скважин с данными")
        
        # Статистика по параметрам
        if well_count > 0:
            for well in list(model_data['well_data'].keys())[:3]:  # Первые 3 скважины для примера
                params = model_data['well_data'][well].keys()
                print(f"    {well}: параметры - {list(params)}")
    
    return models_data


def interpolate_model_data_for_well(model_data, well_name, well_historical_dates) -> dict:
    """
    Интерполировать данные модели для конкретной скважины к её историческим датам
    """
    if well_name not in model_data.get('well_data', {}):
        print(f"Предупреждение: скважина {well_name} не найдена в данных модели")
        return {}
    
    if not well_historical_dates:
        print(f"Предупреждение: нет исторических дат для скважины {well_name}")
        return {}
    
    well_data = model_data['well_data'][well_name]
    model_dates = model_data['dates']
    
    interpolated_well_data = {}
    
    for param, values in well_data.items():
        if len(values) != len(model_dates):
            print(f"Предупреждение: несоответствие размеров для скважины {well_name}, параметр {param}")
            interpolated_well_data[param] = [np.nan] * len(well_historical_dates)
            continue
        
        # Создаем временной ряд для модели
        param_series = pd.Series(values, index=model_dates)
        
        # Реиндексируем на исторические даты этой скважины
        param_series_reindexed = param_series.reindex(well_historical_dates)
        
        # Интерполируем с использованием временной интерполяции
        try:
            # Метод 'time' для временной интерполяции
            param_series_interpolated = param_series_reindexed.interpolate(
                method='time',
                limit_direction='both'
            )
            
            # Заполняем оставшиеся NaN
            param_series_interpolated = param_series_interpolated.fillna(method='ffill').fillna(method='bfill')
            
        except Exception as e:
            print(f"Ошибка интерполяции для скважины {well_name}, параметр {param}: {e}")
            # Используем линейную интерполяцию как запасной вариант
            param_series_interpolated = param_series_reindexed.interpolate(
                method='linear',
                limit_direction='both'
            ).fillna(method='ffill').fillna(method='bfill')
        
        interpolated_well_data[param] = param_series_interpolated.tolist()
    
    return interpolated_well_data


def interpolate_all_models_to_historical_dates(models_data, well_historical_dates_dict):
    """
    Интерполировать данные всех моделей к историческим датам для каждой скважины
    """
    interpolated_models = {}
    
    for model_name, model_info in models_data.items():
        print(f"Интерполяция модели {model_name}...")
        
        interpolated_models[model_name] = {
            'dates': {},  # Даты по скважинам
            'well_data': {}
        }
        
        for well_name, well_historical_dates in well_historical_dates_dict.items():
            # Интерполируем данные для этой скважины
            interpolated_well_data = interpolate_model_data_for_well(
                model_info, well_name, well_historical_dates
            )
            
            if interpolated_well_data:  # Если есть данные
                interpolated_models[model_name]['dates'][well_name] = well_historical_dates
                interpolated_models[model_name]['well_data'][well_name] = interpolated_well_data
    
    return interpolated_models


def create_combined_dataframe_per_well(models_interpolated, historical_df,
                                     well_column='well_fact',
                                     date_column='date_fact',
                                     pressure_column='wpb_bar_fact'):
    """
    Создать объединенный DataFrame с данными, сгруппированными по скважинам
    """
    # Преобразуем исторические даты
    historical_df = historical_df.copy()
    historical_df[date_column] = pd.to_datetime(historical_df[date_column])
    
    well_dataframes = {}
    
    # Сначала получаем список всех скважин
    all_wells = set(historical_df[well_column].unique())
    for model_data in models_interpolated.values():
        all_wells.update(model_data.get('well_data', {}).keys())
    
    print(f"  Всего уникальных скважин: {len(all_wells)}")
    
    for well in all_wells:
        all_records = []
        
        # 1. Добавляем модельные данные для этой скважины
        for model_name, model_info in models_interpolated.items():
            if well in model_info.get('well_data', {}):
                well_dates = model_info['dates'].get(well, [])
                well_data = model_info['well_data'][well]
                
                if not well_dates or not well_data:
                    continue
                
                # Для каждого параметра
                for param, values in well_data.items():
                    if len(values) != len(well_dates):
                        continue
                    
                    # Сопоставляем имена параметров
                    if param == 'wbp':
                        param_display = 'pressure'
                    elif param == 'wgpr':
                        param_display = 'gas_rate'
                    elif param == 'wgir':
                        param_display = 'gas_injection'
                    else:
                        param_display = param
                    
                    # Добавляем записи
                    for date, value in zip(well_dates, values):
                        all_records.append({
                            'date': date,
                            'model': model_name,
                            'parameter': param_display,
                            'value': value
                        })
        
        # 2. Добавляем исторические данные для этой скважины
        well_historical = historical_df[historical_df[well_column] == well]
        
        # Давление
        for _, row in well_historical.iterrows():
            if pd.notna(row.get(pressure_column)):
                all_records.append({
                    'date': row[date_column],
                    'model': 'HISTORICAL',
                    'parameter': 'pressure',
                    'value': row[pressure_column]
                })
        
        # Создаем DataFrame для скважины
        if all_records:
            df_well = pd.DataFrame(all_records)
            df_well = df_well.sort_values(['date', 'parameter', 'model']).reset_index(drop=True)
            well_dataframes[well] = df_well
    
    return well_dataframes


def get_unified_data_per_well(model_names, historical_df,
                            well_column='well_fact',
                            date_column='date_fact',
                            pressure_column='wpb_bar_fact'):
    """
    Основная функция: получить унифицированные данные, сгруппированные по скважинам
    """
    print("=" * 60)
    print("ПОЛУЧЕНИЕ УНИФИЦИРОВАННЫХ ДАННЫХ ПО СКВАЖИНАМ")
    print("=" * 60)
    
    # 1. Получаем исторические даты для каждой скважины
    print("1. Анализ исторических дат по скважинам...")
    well_historical_dates = get_well_historical_dates(
        historical_df, well_column, date_column
    )
    
    if not well_historical_dates:
        print("Ошибка: не удалось получить исторические даты")
        return {}, {}
    
    # 2. Получаем список скважин
    well_names = list(well_historical_dates.keys())
    print(f"Скважины для анализа: {well_names}")
    
    # 3. Загружаем сырые данные из моделей
    print("\n2. Загрузка сырых данных из моделей...")
    models_raw = get_raw_model_data(model_names, well_names, ['wbp', 'wgpr', 'wgir'])
    
    if not models_raw:
        print("Ошибка: не удалось загрузить данные моделей")
        return {}, {}
    
    print(f"Загружены данные из {len(models_raw)} моделей")
    
    # 4. Интерполируем данные моделей к историческим датам каждой скважины
    print("\n3. Интерполяция данных моделей к историческим датам...")
    models_interpolated = interpolate_all_models_to_historical_dates(
        models_raw, well_historical_dates
    )
    
    # 5. Создаем объединенные DataFrame по скважинам
    print("\n4. Создание объединенных DataFrame по скважинам...")
    well_dataframes = create_combined_dataframe_per_well(
        models_interpolated, historical_df,
        well_column=well_column,
        date_column=date_column,
        pressure_column=pressure_column
    )
    
    print(f"Созданы DataFrame для {len(well_dataframes)} скважин")
    
    # 6. Выводим информацию о загруженных параметрах
    print("\n5. Анализ загруженных параметров...")
    param_stats = {}
    for well, df_well in well_dataframes.items():
        params = df_well['parameter'].unique()
        for param in params:
            if param not in param_stats:
                param_stats[param] = set()
            param_stats[param].add(well)
    
    print("Загружены следующие параметры:")
    for param, wells in param_stats.items():
        print(f"  {param}: {len(wells)} скважин")
    
    return well_dataframes, models_interpolated


def save_to_excel_structured(well_dataframes, historical_df, models_interpolated, 
                           output_path="structured_comparison.xlsx"):
    """
    Сохранить данные в Excel файл с указанной структурой:
    1 строка: well, date, wbp_hist, model_name1, "", "", model_name2, "", "", ...
    2 строка: "", "", "", wbp_model, wgpr_model, wgir_model, wbp_model, wgpr_model, wgir_model...
    А ниже соответствующие данные.
    
    Parameters:
    -----------
    well_dataframes : dict
        DataFrame по скважинам
    historical_df : pd.DataFrame
        Исторические данные
    models_interpolated : dict
        Интерполированные данные моделей
    output_path : str
        Путь для сохранения Excel файла
    """
    print(f"\nСохранение данных в Excel файл: {output_path}")
    
    # Создаем новый Workbook
    wb = Workbook()
    
    # Для каждой скважины создаем отдельный лист
    for well_name, df_well in well_dataframes.items():
        print(f"  Обработка скважины {well_name}...")
        
        # Создаем новый лист с именем скважины (ограничиваем длину имени листа)
        sheet_name = str(well_name)[:31]  # Excel ограничивает имена листов 31 символом
        if well_name == list(well_dataframes.keys())[0]:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Получаем уникальные модели для этой скважины (без HISTORICAL)
        models_in_well = sorted([m for m in df_well['model'].unique() if m != 'HISTORICAL'])
        
        if not models_in_well:
            print(f"    Предупреждение: для скважины {well_name} нет модельных данных")
            ws.append(["Нет модельных данных для этой скважины"])
            continue
        
        # Получаем уникальные даты для этой скважины
        well_dates = sorted(df_well['date'].unique())
        
        if not well_dates:
            print(f"    Предупреждение: для скважины {well_name} нет данных по датам")
            ws.append(["Нет данных по датам для этой скважины"])
            continue
        
        # === СТРОКА 1: Заголовки ===
        header_row1 = ['well', 'date', 'wbp_hist']
        for model in models_in_well:
            header_row1.append(model)  # Название модели
            header_row1.append("")      # Пустой столбец
            header_row1.append("")      # Пустой столбец
        
        ws.append(header_row1)
        
        # === СТРОКА 2: Подзаголовки параметров ===
        header_row2 = ['', '', '']
        for model in models_in_well:
            header_row2.append('wbp_model')
            header_row2.append('wgpr_model')
            header_row2.append('wgir_model')
        
        ws.append(header_row2)
        
        # === ДАННЫЕ ПО СТРОКАМ ===
        for date in well_dates:
            # Преобразуем дату в строку (обрабатываем разные форматы дат)
            try:
                if isinstance(date, (pd.Timestamp, datetime)):
                    date_str = date.strftime('%Y-%m-%d')
                elif isinstance(date, np.datetime64):
                    # Преобразуем numpy.datetime64 в pandas.Timestamp
                    date_str = pd.Timestamp(date).strftime('%Y-%m-%d')
                else:
                    date_str = str(date)
            except Exception as e:
                print(f"    Ошибка преобразования даты {date}: {e}")
                date_str = str(date)
            
            # Получаем историческое давление для этой даты
            hist_pressure = None
            hist_row = df_well[(df_well['date'] == date) & 
                              (df_well['model'] == 'HISTORICAL') & 
                              (df_well['parameter'] == 'pressure')]
            if not hist_row.empty:
                hist_pressure = hist_row['value'].iloc[0]
            
            # Создаем строку данных
            data_row = [well_name, date_str, hist_pressure]
            
            # Добавляем модельные данные для каждой модели
            for model in models_in_well:
                # Давление (wbp)
                wbp_value = None
                wbp_row = df_well[(df_well['date'] == date) & 
                                 (df_well['model'] == model) & 
                                 (df_well['parameter'] == 'pressure')]
                if not wbp_row.empty:
                    wbp_value = wbp_row['value'].iloc[0]
                
                # Добыча газа (wgpr)
                wgpr_value = None
                wgpr_row = df_well[(df_well['date'] == date) & 
                                  (df_well['model'] == model) & 
                                  (df_well['parameter'] == 'gas_rate')]
                if not wgpr_row.empty:
                    wgpr_value = wgpr_row['value'].iloc[0]
                
                # Закачка газа (wgir)
                wgir_value = None
                wgir_row = df_well[(df_well['date'] == date) & 
                                  (df_well['model'] == model) & 
                                  (df_well['parameter'] == 'gas_injection')]
                if not wgir_row.empty:
                    wgir_value = wgir_row['value'].iloc[0]
                
                data_row.extend([wbp_value, wgpr_value, wgir_value])
            
            ws.append(data_row)
        
        # Автонастройка ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"    ✓ Данные для скважины {well_name} сохранены ({len(well_dates)} строк)")
    
    # Сохраняем файл
    try:
        wb.save(output_path)
        print(f"\n✓ Файл успешно сохранен: {output_path}")
        print(f"  Всего листов: {len(wb.sheetnames)} (по одному на скважину)")
        
        # Выводим информацию о структуре файла
        if wb.sheetnames:
            first_sheet = wb[wb.sheetnames[0]]
            models_in_first_well = sorted([m for m in well_dataframes[list(well_dataframes.keys())[0]]['model'].unique() if m != 'HISTORICAL'])
            print(f"  Структура данных в листе '{wb.sheetnames[0]}':")
            print(f"    - Столбцов: {first_sheet.max_column}")
            print(f"    - Строк данных: {first_sheet.max_row - 2}")  # минус 2 строки заголовков
            print(f"    - Моделей: {len(models_in_first_well)}")
            
    except Exception as e:
        print(f"✗ Ошибка при сохранении файла: {e}")
        import traceback
        traceback.print_exc()
    
    return output_path

def main():
    """
    Основная функция для получения унифицированных данных
    """
    print("=" * 80)
    print("ПОЛУЧЕНИЕ УНИФИЦИРОВАННЫХ ДАННЫХ ДЛЯ СРАВНЕНИЯ")
    print("=" * 80)
    
    try:
        # 1. Читаем фактические данные из файла
        print(f"\n1. Чтение фактических данных из файла: {WBP_FACT_TXT}")
        fact_file_path = os.path.join(PROJECT_FOLDER_PATH, WBP_FACT_TXT)
        
        if not os.path.exists(fact_file_path):
            print(f"ОШИБКА: Файл {fact_file_path} не найден!")
            print(f"Текущая директория проекта: {PROJECT_FOLDER_PATH}")
            return None
        
        df_fact = parse_fact_well_data(fact_file_path)
        
        if df_fact.empty:
            print("ОШИБКА: Не удалось загрузить фактические данные!")
            return None
        
        print(f"Загружено {len(df_fact)} записей фактических данных")
        print(f"Уникальных скважин: {df_fact['well_fact'].nunique()}")
        
        # 2. Получаем унифицированные данные
        print(f"\n2. Получение унифицированных данных...")
        well_dataframes, models_interpolated = get_unified_data_per_well(
            model_names=MODEL_NAMES,
            historical_df=df_fact,
            well_column='well_fact',
            date_column='date_fact',
            pressure_column='wpb_bar_fact'
        )
        
        if not well_dataframes:
            print("ОШИБКА: Не удалось получить унифицированные данные!")
            return None
        
        # 3. Выводим статистику
        print("\n" + "=" * 60)
        print("РЕЗУЛЬТАТЫ")
        print("=" * 60)
        
        total_records = 0
        param_counts = {}
        
        for well, df_well in well_dataframes.items():
            records = len(df_well)
            total_records += records
            
            # Считаем параметры
            for param in df_well['parameter'].unique():
                if param not in param_counts:
                    param_counts[param] = 0
                param_counts[param] += len(df_well[df_well['parameter'] == param])
            
            print(f"Скважина {well}: {records} записей, "
                  f"{df_well['parameter'].nunique()} параметров, "
                  f"{df_well['date'].nunique()} дат")
        
        print(f"\nВсего записей: {total_records}")
        print(f"Всего скважин: {len(well_dataframes)}")
        print(f"Распределение по параметрам:")
        for param, count in param_counts.items():
            print(f"  {param}: {count} записей")
        
        # 4. Сохраняем данные в Excel с указанной структурой
        print(f"\n3. Сохранение данных в Excel файл...")
        output_excel = os.path.join(PROJECT_FOLDER_PATH, "structured_comparison.xlsx")
        save_to_excel_structured(well_dataframes, df_fact, models_interpolated, output_excel)
        
        # Возвращаем данные для дальнейшего использования
        return {
            'fact_data': df_fact,
            'well_dataframes': well_dataframes,
            'models_interpolated': models_interpolated,
            'excel_file': output_excel,
            'statistics': {
                'total_records': total_records,
                'total_wells': len(well_dataframes),
                'parameter_counts': param_counts
            }
        }
        
    except Exception as e:
        print(f"\nОШИБКА ВО ВРЕМЯ ВЫПОЛНЕНИЯ: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# БЛОК ВЫПОЛНЕНИЯ СКРИПТА
result_data = main()
if result_data:
    print("\n" + "=" * 80)
    print("ДАННЫЕ УСПЕШНО ПОДГОТОВЛЕНЫ ДЛЯ СРАВНЕНИЯ!")
    print("=" * 80)
    print(f"Структура данных содержит:")
    print(f"  - Фактические данные: {len(result_data['fact_data'])} записей")
    print(f"  - Данные по скважинам: {len(result_data['well_dataframes'])} скважин")
    print(f"  - Интерполированные модельные данные: {len(result_data['models_interpolated'])} моделей")
    print(f"  - Всего записей: {result_data['statistics']['total_records']}")
    print(f"  - Параметры: {list(result_data['statistics']['parameter_counts'].keys())}")
    print(f"  - Excel файл: {result_data['excel_file']}")
    
    # Пример доступа к данным
    if result_data['well_dataframes']:
        first_well = list(result_data['well_dataframes'].keys())[0]
        print(f"\nПример данных для скважины {first_well}:")
        print(result_data['well_dataframes'][first_well].head())